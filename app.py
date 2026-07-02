"""
Gestão de Pagamentos por Conta 2026 — SERVE
Fluxo completo: login por gestor -> importar clientes -> calcular PPC -> associar guias
-> gerar e enviar emails -> exportar.

Base de dados: Supabase (Postgres), com Row Level Security — cada gestor só vê e edita
os clientes que lhe estão atribuídos (campo Gestor_Email); o admin vê tudo.

Correr com:  streamlit run app.py

Configuração necessária em .streamlit/secrets.toml (local) ou em
Settings → Secrets (Streamlit Community Cloud):

    SUPABASE_URL = "https://XXXXXXXX.supabase.co"
    SUPABASE_ANON_KEY = "a chave 'anon public' do projeto"
    SUPABASE_SERVICE_KEY = "a chave 'service_role' do projeto"   # opcional, só o admin precisa

Ver GUIA_SUPABASE.md para o processo completo passo a passo.
"""

import io
import math
import re
import smtplib
import ssl
from datetime import date, datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from supabase import Client, create_client

# ---------------------------------------------------------------------------
# Configuração geral
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Pagamentos por Conta 2026", layout="wide", page_icon="💶")

CLIENT_COLS = [
    "NIF", "Nome", "Email",
    "Gestor_Nome", "Gestor_Email",
    "Volume_2025", "Coleta_2025", "Retencoes_2025",
    "Guia1_Emitida", "Guia2_Emitida", "Guia3_Emitida",
    "Email1_Enviado", "Email2_Enviado", "Email3_Enviado",
    "Notas",
]
BOOL_COLS = [c for c in CLIENT_COLS if c.startswith("Guia") or c.startswith("Email") and c.endswith(("Emitida", "Enviado"))]
TEXT_COLS = ["NIF", "Nome", "Email", "Gestor_Nome", "Gestor_Email", "Notas"]

# Mapeamento entre os nomes de colunas usados na app (iguais aos de sempre) e os
# nomes das colunas na base de dados Postgres (o Postgres normaliza tudo para
# minúsculas, por isso a tabela "clientes" no Supabase usa nomes em minúsculas).
COLUMN_MAP_TO_DB = {
    "NIF": "nif", "Nome": "nome", "Email": "email",
    "Gestor_Nome": "gestor_nome", "Gestor_Email": "gestor_email",
    "Volume_2025": "volume_2025", "Coleta_2025": "coleta_2025", "Retencoes_2025": "retencoes_2025",
    "Guia1_Emitida": "guia1_emitida", "Guia2_Emitida": "guia2_emitida", "Guia3_Emitida": "guia3_emitida",
    "Email1_Enviado": "email1_enviado", "Email2_Enviado": "email2_enviado", "Email3_Enviado": "email3_enviado",
    "Notas": "notas",
}
COLUMN_MAP_FROM_DB = {v: k for k, v in COLUMN_MAP_TO_DB.items()}

DEFAULT_TEMPLATES = {
    1: {
        "assunto": "Pagamentos por Conta 2026 — {nome}",
        "corpo": (
            "Exmo(a). Sr(a).,\n\n"
            "No seguimento do apuramento do IRC referente a 2025, informamos que a {nome} "
            "(NIF {nif}) tem pagamentos por conta a efetuar em 2026, nos seguintes montantes e prazos:\n\n"
            "• 1.º Pagamento por Conta: {pag1} € — até {data1}\n"
            "• 2.º Pagamento por Conta: {pag2} € — até {data2}\n"
            "• 3.º Pagamento por Conta: {pag3} € — até {data3}\n\n"
            "Total anual: {total} €\n\n"
            "Segue em anexo a guia referente ao 1.º pagamento. Solicitamos que proceda ao pagamento até à data "
            "indicada, de forma a evitar juros de mora.\n\n"
            "As guias do 2.º e 3.º pagamento serão enviadas atempadamente.\n\n"
            "Ficamos ao dispor para qualquer esclarecimento.\n\n"
            "Com os melhores cumprimentos,"
        ),
    },
    2: {
        "assunto": "2.º Pagamento por Conta 2026 — {nome}",
        "corpo": (
            "Exmo(a). Sr(a).,\n\n"
            "No seguimento do 1.º pagamento por conta já efetuado, relembramos que o 2.º pagamento por conta "
            "da {nome} (NIF {nif}) vence a {data2}, no valor de {pag2} €.\n\n"
            "Segue em anexo a respetiva guia.\n\n"
            "Ficamos ao dispor para qualquer esclarecimento.\n\n"
            "Com os melhores cumprimentos,"
        ),
    },
    3: {
        "assunto": "3.º Pagamento por Conta 2026 — {nome}",
        "corpo": (
            "Exmo(a). Sr(a).,\n\n"
            "No seguimento dos pagamentos por conta já efetuados, relembramos que o 3.º e último pagamento por "
            "conta da {nome} (NIF {nif}) vence a {data3}, no valor de {pag3} €.\n\n"
            "Segue em anexo a respetiva guia.\n\n"
            "Ficamos ao dispor para qualquer esclarecimento.\n\n"
            "Com os melhores cumprimentos,"
        ),
    },
}

# ---------------------------------------------------------------------------
# Ligação ao Supabase
# ---------------------------------------------------------------------------
SUPABASE_URL = st.secrets.get("SUPABASE_URL")
SUPABASE_ANON_KEY = st.secrets.get("SUPABASE_ANON_KEY")
SUPABASE_SERVICE_KEY = st.secrets.get("SUPABASE_SERVICE_KEY")  # só necessário para o admin criar gestores


def get_client() -> Client:
    """Devolve o cliente Supabase desta sessão de browser (um por utilizador —
    nunca partilhado entre utilizadores, para que a sessão de autenticação de
    cada um não se misture com a de outro)."""
    if "sb_client" not in st.session_state:
        st.session_state.sb_client = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
    return st.session_state.sb_client


def get_admin_client() -> Client | None:
    """Cliente com a chave 'service_role' — só usado nas ações de administração
    de contas (criar gestores). Nunca deve ser usado para ler/escrever clientes."""
    if not SUPABASE_SERVICE_KEY:
        return None
    return create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)


# ---------------------------------------------------------------------------
# Autenticação e perfil (admin vs. gestor)
# ---------------------------------------------------------------------------
def carregar_perfil():
    client = get_client()
    user_id = st.session_state.user.id
    resp = client.table("perfis").select("*").eq("id", user_id).execute()
    if resp.data:
        st.session_state.perfil = resp.data[0]
    else:
        # Ainda não existe perfil (não devia acontecer, o trigger cria-o) — usa um por omissão seguro.
        st.session_state.perfil = {"id": user_id, "email": st.session_state.user.email, "nome": st.session_state.user.email, "role": "gestor"}


def ecra_login():
    st.title("🔒 Gestão de Pagamentos por Conta 2026")
    st.caption("SERVE — Contabilidade e Viabilização Empresarial. Inicia sessão com a tua conta de gestor.")

    if not SUPABASE_URL or not SUPABASE_ANON_KEY:
        st.error(
            "A app ainda não está ligada ao Supabase. Falta configurar SUPABASE_URL e "
            "SUPABASE_ANON_KEY em Settings → Secrets. Ver GUIA_SUPABASE.md."
        )
        st.stop()

    email = st.text_input("Email")
    pwd = st.text_input("Password", type="password")
    if st.button("Entrar", type="primary"):
        try:
            client = get_client()
            res = client.auth.sign_in_with_password({"email": email, "password": pwd})
            st.session_state.user = res.user
            carregar_perfil()
            st.rerun()
        except Exception:
            st.error("Email ou password incorretos.")


def sidebar_utilizador():
    perfil = st.session_state.perfil
    with st.sidebar:
        papel = "Administrador" if perfil["role"] == "admin" else "Gestor"
        st.success(f"👤 {perfil['nome'] or perfil['email']}  \n**{papel}**")
        if st.button("Sair"):
            try:
                get_client().auth.sign_out()
            except Exception:
                pass
            for k in ("user", "perfil", "sb_client"):
                st.session_state.pop(k, None)
            st.rerun()
        st.divider()


if "user" not in st.session_state:
    ecra_login()
    st.stop()

sou_admin = st.session_state.perfil["role"] == "admin"
meu_email_login = st.session_state.perfil["email"]

# ---------------------------------------------------------------------------
# Persistência (Supabase — a RLS garante que cada gestor só vê/edita a sua carteira)
# ---------------------------------------------------------------------------
def carregar_clientes_db() -> pd.DataFrame:
    client = get_client()
    resp = client.table("clientes").select("*").execute()
    rows = resp.data or []
    if not rows:
        return pd.DataFrame(columns=CLIENT_COLS)
    df = pd.DataFrame(rows).rename(columns=COLUMN_MAP_FROM_DB)
    for c in CLIENT_COLS:
        if c not in df.columns:
            df[c] = False if c in BOOL_COLS else ("" if c in TEXT_COLS else 0.0)
    for c in BOOL_COLS:
        df[c] = df[c].fillna(False).astype(bool)
    return df[CLIENT_COLS]


def guardar_clientes_db(df: pd.DataFrame):
    """Substitui os clientes visíveis por este utilizador pelo conteúdo de df.
    Um gestor só vê/apaga/insere os seus próprios clientes (RLS trata do âmbito);
    o admin substitui a lista completa, tal como sempre funcionou."""
    client = get_client()
    df2 = clean_df(df).copy()

    if not sou_admin:
        # Um gestor só pode gravar clientes atribuídos a si mesmo — evita
        # rejeições da RLS e garante que carteiras novas ficam bem atribuídas.
        df2["Gestor_Email"] = meu_email_login
        if not perfil_nome_vazio():
            df2["Gestor_Nome"] = st.session_state.perfil["nome"]

    client.table("clientes").delete().neq("nif", "").execute()
    if not df2.empty:
        registos = df2.rename(columns=COLUMN_MAP_TO_DB).to_dict("records")
        client.table("clientes").upsert(registos, on_conflict="nif").execute()


def perfil_nome_vazio() -> bool:
    return not (st.session_state.perfil.get("nome") or "").strip()


def persistir_clientes(df: pd.DataFrame):
    """Atualiza a sessão E grava imediatamente no Supabase."""
    df = clean_df(df)
    if not sou_admin:
        df["Gestor_Email"] = meu_email_login
        if not perfil_nome_vazio():
            df["Gestor_Nome"] = st.session_state.perfil["nome"]
    st.session_state.clientes = df
    guardar_clientes_db(df)


def carregar_log_db() -> list:
    client = get_client()
    resp = client.table("log_envios").select("data, nif, nome, pagamento, estado").order("id").execute()
    return resp.data or []


def guardar_log_entry_db(entry: dict):
    client = get_client()
    client.table("log_envios").insert(entry).execute()


def registar_log(entry: dict):
    st.session_state.log_envio.append(entry)
    guardar_log_entry_db(entry)


def carregar_config_db():
    client = get_client()
    resp = client.table("config").select("params_json, templates_json").eq("id", 1).execute()
    if not resp.data:
        return None, None
    row = resp.data[0]
    params_json, templates_json = row.get("params_json"), row.get("templates_json")
    params, templates = None, None
    if params_json:
        params = {
            "limiar_volume": params_json["limiar_volume"], "taxa_baixa": params_json["taxa_baixa"],
            "taxa_alta": params_json["taxa_alta"], "limite_dispensa": params_json["limite_dispensa"],
            "data1": date.fromisoformat(params_json["data1"]), "data2": date.fromisoformat(params_json["data2"]),
            "data3": date.fromisoformat(params_json["data3"]),
        }
    if templates_json:
        templates = {int(k): v for k, v in templates_json.items()}
    return params, templates


def guardar_config_db(params: dict, templates: dict):
    """Só o admin tem permissão (RLS) para escrever a configuração global."""
    if not sou_admin:
        return
    params_serializ = dict(params)
    for k in ("data1", "data2", "data3"):
        params_serializ[k] = params[k].isoformat()
    templates_serializ = {str(k): v for k, v in templates.items()}
    client = get_client()
    client.table("config").upsert(
        {"id": 1, "params_json": params_serializ, "templates_json": templates_serializ}
    ).execute()


# ---------------------------------------------------------------------------
# Estado
# ---------------------------------------------------------------------------
def init_state():
    if "perfil" not in st.session_state:
        carregar_perfil()
    if "clientes" not in st.session_state:
        st.session_state.clientes = carregar_clientes_db()
    if "guias" not in st.session_state:
        st.session_state.guias = {}  # {(nif, n_pagamento): (filename, bytes)} — só dura a sessão atual
    if "params" not in st.session_state or "templates" not in st.session_state:
        params_db, templates_db = carregar_config_db()
        if "params" not in st.session_state:
            st.session_state.params = params_db or {
                "limiar_volume": 500000.0,
                "taxa_baixa": 0.80,
                "taxa_alta": 0.95,
                "limite_dispensa": 200.0,
                "data1": date(2026, 7, 31),
                "data2": date(2026, 9, 30),
                "data3": date(2026, 12, 15),
            }
        if "templates" not in st.session_state:
            st.session_state.templates = templates_db or {k: v.copy() for k, v in DEFAULT_TEMPLATES.items()}
    if "log_envio" not in st.session_state:
        st.session_state.log_envio = carregar_log_db()


def parse_numero_pt(serie: pd.Series) -> pd.Series:
    """Converte números em formato português (vírgula decimal, ponto de milhares) ou internacional para float."""
    def conv(v):
        if pd.isna(v):
            return None
        s = str(v).strip()
        if s == "":
            return None
        s = s.replace(" ", "").replace("€", "")
        if "," in s and "." in s:
            # ex: "1.234,56" -> milhares '.', decimal ','
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            # ex: "4870,84" -> decimal ','
            s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return serie.apply(conv)


def ler_ficheiro_importacao(uploaded_file) -> pd.DataFrame:
    """Lê CSV ou Excel de forma tolerante: deteta o encoding do CSV automaticamente
    (ficheiros exportados do Excel em português costumam vir em Windows-1252/ISO-8859-1,
    não UTF-8) e trata números com vírgula decimal."""
    nome = uploaded_file.name.lower()
    if nome.endswith(".csv"):
        raw = uploaded_file.read()
        texto = None
        for enc in ("utf-8-sig", "utf-8", "cp1252", "iso-8859-1"):
            try:
                texto = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if texto is None:
            texto = raw.decode("latin1", errors="replace")
        df = pd.read_csv(io.StringIO(texto), sep=None, engine="python")
    else:
        df = pd.read_excel(uploaded_file)

    df.columns = [str(c).strip() for c in df.columns]
    for c in ("Volume_2025", "Coleta_2025", "Retencoes_2025"):
        if c in df.columns:
            df[c] = parse_numero_pt(df[c])
    if "Email" in df.columns:
        df["Email"] = df["Email"].astype(str).str.strip().str.rstrip(",").str.strip()
        df.loc[df["Email"] == "nan", "Email"] = ""
    if "NIF" in df.columns:
        df["NIF"] = df["NIF"].astype(str).str.strip()
    return df


def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in CLIENT_COLS:
        if c not in df.columns:
            df[c] = False if c in BOOL_COLS else ("" if c in TEXT_COLS else 0.0)
    for c in ("Volume_2025", "Coleta_2025", "Retencoes_2025"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    for c in BOOL_COLS:
        df[c] = df[c].fillna(False).astype(bool)
    for c in TEXT_COLS:
        df[c] = df[c].fillna("").astype(str).str.strip()
    return df[CLIENT_COLS]


# ---------------------------------------------------------------------------
# Cálculo PPC
# ---------------------------------------------------------------------------
def calcular_ppc(df: pd.DataFrame, params: dict) -> pd.DataFrame:
    df = df.copy()
    df["Base_Calculo"] = (df["Coleta_2025"] - df["Retencoes_2025"]).clip(lower=0)
    df["Taxa"] = df["Volume_2025"].apply(
        lambda v: params["taxa_baixa"] if v <= params["limiar_volume"] else params["taxa_alta"]
    )
    df["Dispensado"] = df["Base_Calculo"] < params["limite_dispensa"]
    df["Total_PPC_Base"] = df["Base_Calculo"] * df["Taxa"]  # valor de referência, antes de repartir

    def parcela(row):
        if row["Dispensado"] or row["Total_PPC_Base"] <= 0:
            return 0
        # art. 105.º n.º 2/3 CIRC: "repartido por três montantes iguais, arredondados, por excesso, para euros"
        return math.ceil(round(row["Total_PPC_Base"] / 3, 6) - 1e-9)

    df["Pag1"] = df.apply(parcela, axis=1)
    df["Pag2"] = df["Pag1"]
    df["Pag3"] = df["Pag1"]
    df["Total_PPC"] = df["Pag1"] + df["Pag2"] + df["Pag3"]  # valor efetivamente cobrado (após arredondamentos)
    return df


# ---------------------------------------------------------------------------
# Export Excel (mesma formatação da folha de controlo entregue anteriormente)
# ---------------------------------------------------------------------------
def gerar_excel(df_calc: pd.DataFrame, params: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Controlo PPC 2026"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    FONT = "Arial"
    HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    DISPENSA_FILL = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    HEADER_FONT = Font(name=FONT, color="FFFFFF", bold=True, size=10)
    TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
    BLACK = Font(name=FONT, color="000000")
    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Controlo de Pagamentos por Conta — 2026"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    headers = [
        "NIF", "Nome", "Email", "Gestor (nome)", "Gestor (email)", "Volume 2025", "Coleta 2025", "Retenções 2025",
        "Base de Cálculo", "Taxa", "Total PPC", "Dispensado",
        "1º Pagamento", "2º Pagamento", "3º Pagamento",
        "Data Limite 1º", "Data Limite 2º", "Data Limite 3º",
        "Guia1 Emitida", "Guia2 Emitida", "Guia3 Emitida",
        "Email1 Enviado", "Email2 Enviado", "Email3 Enviado", "Notas",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[4].height = 30

    widths = [12, 26, 24, 18, 24, 13, 13, 13, 13, 8, 12, 11, 11, 11, 11, 12, 12, 12, 11, 11, 11, 11, 11, 11, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 5
    for _, r in df_calc.iterrows():
        vals = [
            r["NIF"], r["Nome"], r["Email"], r["Gestor_Nome"], r["Gestor_Email"],
            r["Volume_2025"], r["Coleta_2025"], r["Retencoes_2025"],
            r["Base_Calculo"], r["Taxa"], r["Total_PPC"], "Sim" if r["Dispensado"] else "Não",
            r["Pag1"], r["Pag2"], r["Pag3"],
            params["data1"] if not r["Dispensado"] else None,
            params["data2"] if not r["Dispensado"] else None,
            params["data3"] if not r["Dispensado"] else None,
            "Sim" if r["Guia1_Emitida"] else "Não",
            "Sim" if r["Guia2_Emitida"] else "Não",
            "Sim" if r["Guia3_Emitida"] else "Não",
            "Sim" if r["Email1_Enviado"] else "Não",
            "Sim" if r["Email2_Enviado"] else "Não",
            "Sim" if r["Email3_Enviado"] else "Não",
            r["Notas"],
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = BLACK
            c.border = BORDER
            if i in (6, 7, 8, 9, 11, 13, 14, 15):
                c.number_format = "#,##0.00"
            if i == 10:
                c.number_format = "0%"
            if i in (16, 17, 18):
                c.number_format = "dd/mm/yyyy"
        if r["Dispensado"]:
            for i in range(1, len(vals) + 1):
                ws.cell(row=row, column=i).fill = DISPENSA_FILL
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Emails
# ---------------------------------------------------------------------------
def formatar_valor(v):
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def render_template(template: dict, row: pd.Series, params: dict) -> tuple[str, str]:
    ctx = {
        "nome": row["Nome"],
        "nif": row["NIF"],
        "email": row["Email"],
        "pag1": formatar_valor(row["Pag1"]),
        "pag2": formatar_valor(row["Pag2"]),
        "pag3": formatar_valor(row["Pag3"]),
        "total": formatar_valor(row["Total_PPC"]),
        "data1": params["data1"].strftime("%d/%m/%Y"),
        "data2": params["data2"].strftime("%d/%m/%Y"),
        "data3": params["data3"].strftime("%d/%m/%Y"),
    }
    assunto = template["assunto"].format(**ctx)
    corpo = template["corpo"].format(**ctx)
    return assunto, corpo


def enviar_email(smtp_cfg, destinatario, assunto, corpo, anexos, cc=None):
    cc_list = [e.strip() for e in (cc or []) if e and e.strip()]

    msg = MIMEMultipart()
    msg["From"] = smtp_cfg["remetente"]
    msg["To"] = destinatario
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    msg["Subject"] = assunto
    msg.attach(MIMEText(corpo, "plain", "utf-8"))
    for filename, filebytes in anexos:
        part = MIMEApplication(filebytes, Name=filename)
        part["Content-Disposition"] = f'attachment; filename="{filename}"'
        msg.attach(part)

    todos_destinatarios = [destinatario] + cc_list

    context = ssl.create_default_context()
    if smtp_cfg["tls"]:
        with smtplib.SMTP(smtp_cfg["host"], smtp_cfg["porta"], timeout=30) as server:
            server.starttls(context=context)
            server.login(smtp_cfg["utilizador"], smtp_cfg["password"])
            server.sendmail(smtp_cfg["remetente"], todos_destinatarios, msg.as_string())
    else:
        with smtplib.SMTP_SSL(smtp_cfg["host"], smtp_cfg["porta"], context=context, timeout=30) as server:
            server.login(smtp_cfg["utilizador"], smtp_cfg["password"])
            server.sendmail(smtp_cfg["remetente"], todos_destinatarios, msg.as_string())


def extrair_nif_de_filename(filename: str):
    m = re.search(r"\b(\d{9})\b", filename)
    return m.group(1) if m else None


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
init_state()
sidebar_utilizador()

st.title("💶 Gestão de Pagamentos por Conta 2026")
st.caption("SERVE — Contabilidade e Viabilização Empresarial")

with st.sidebar:
    st.header("Parâmetros de Cálculo")
    p = st.session_state.params
    if sou_admin:
        p["limiar_volume"] = st.number_input("Limiar Volume de Negócios (€)", value=float(p["limiar_volume"]), step=10000.0)
        p["taxa_baixa"] = st.number_input("Taxa se Volume ≤ limiar", value=float(p["taxa_baixa"]), step=0.01, format="%.2f")
        p["taxa_alta"] = st.number_input("Taxa se Volume > limiar", value=float(p["taxa_alta"]), step=0.01, format="%.2f")
        p["limite_dispensa"] = st.number_input("Limite de dispensa (€)", value=float(p["limite_dispensa"]), step=10.0)
        st.divider()
        p["data1"] = st.date_input("Data limite 1.º Pagamento", value=p["data1"])
        p["data2"] = st.date_input("Data limite 2.º Pagamento", value=p["data2"])
        p["data3"] = st.date_input("Data limite 3.º Pagamento", value=p["data3"])
    else:
        st.caption("Estes parâmetros são definidos pelo administrador e aplicam-se a todos os clientes.")
        st.write(f"Limiar Volume de Negócios: **{p['limiar_volume']:,.2f} €**")
        st.write(f"Taxa ≤ limiar / > limiar: **{p['taxa_baixa']:.0%} / {p['taxa_alta']:.0%}**")
        st.write(f"Limite de dispensa: **{p['limite_dispensa']:,.2f} €**")
        st.divider()
        st.write(f"Data limite 1.º Pagamento: **{p['data1'].strftime('%d/%m/%Y')}**")
        st.write(f"Data limite 2.º Pagamento: **{p['data2'].strftime('%d/%m/%Y')}**")
        st.write(f"Data limite 3.º Pagamento: **{p['data3'].strftime('%d/%m/%Y')}**")
    st.divider()
    st.caption("Fórmula: Total PPC = (Coleta IRC − Retenções) × Taxa, repartido em 3 prestações iguais, cada uma arredondada por excesso para euro (art. 105.º CIRC). Dispensa se (Coleta − Retenções) < limite definido.")

nomes_tabs = ["📊 Dashboard", "📋 Clientes", "🧮 Cálculo PPC", "📎 Guias", "✉️ Emails", "⬇️ Exportar"]
if sou_admin:
    nomes_tabs.append("👥 Gestores")
tabs = st.tabs(nomes_tabs)
tab_dash, tab1, tab2, tab3, tab4, tab5 = tabs[:6]
tab_gestores = tabs[6] if sou_admin else None

# --- TAB 1: Clientes ---------------------------------------------------
with tab1:
    st.subheader("Importar / Editar Clientes")
    if not sou_admin:
        st.caption(f"Estás a ver apenas os clientes atribuídos a ti ({meu_email_login}).")

    col1, col2 = st.columns([2, 1])
    with col1:
        up = st.file_uploader("Importar CSV ou Excel (colunas: NIF, Nome, Email, Gestor_Nome, Gestor_Email, Volume_2025, Coleta_2025, Retencoes_2025)", type=["csv", "xlsx"])
        if up is not None:
            try:
                novo = ler_ficheiro_importacao(up)
                novo = clean_df(novo)
                modo = st.radio("Modo de importação", ["Substituir tudo", "Adicionar aos existentes"], horizontal=True, key="modo_import")
                if st.button("Confirmar importação"):
                    if modo == "Substituir tudo":
                        persistir_clientes(novo)
                    else:
                        persistir_clientes(
                            clean_df(pd.concat([st.session_state.clientes, novo], ignore_index=True))
                            .drop_duplicates(subset="NIF", keep="last")
                        )
                    st.success(f"{len(novo)} clientes importados e guardados.")
                    st.rerun()
            except Exception as e:
                st.error(f"Erro ao importar: {e}")
    with col2:
        template_csv = pd.DataFrame(
            [{"NIF": "500123456", "Nome": "Empresa Exemplo, Lda.", "Email": "geral@exemplo.pt",
              "Gestor_Nome": "Ana Gestora", "Gestor_Email": "ana@serve.pt",
              "Volume_2025": 10000, "Coleta_2025": 2000, "Retencoes_2025": 200}]
        ).to_csv(index=False, sep=";")
        st.download_button("📥 Template CSV", template_csv, file_name="template_clientes.csv", mime="text/csv")

    st.markdown("**Tabela de clientes** — pode editar diretamente, adicionar ou apagar linhas.")
    if sou_admin:
        st.caption("O 'Gestor' é opcional — se preenchido, o email desse gestor entra automaticamente em CC quando enviares os avisos de pagamento a este cliente, e esse gestor passa a ver este cliente na sua área.")
    else:
        st.caption("Os novos clientes que adicionares aqui ficam automaticamente atribuídos a ti.")
    col_config = {
        "Volume_2025": st.column_config.NumberColumn("Volume Negócios 2025 (campo 411)", format="%.2f"),
        "Coleta_2025": st.column_config.NumberColumn("Coleta IRC 2025 (campo 351)", format="%.2f"),
        "Retencoes_2025": st.column_config.NumberColumn("Retenções 2025 (campo 359)", format="%.2f"),
        "Guia1_Emitida": st.column_config.CheckboxColumn("Guia 1 Emitida"),
        "Guia2_Emitida": st.column_config.CheckboxColumn("Guia 2 Emitida"),
        "Guia3_Emitida": st.column_config.CheckboxColumn("Guia 3 Emitida"),
        "Email1_Enviado": st.column_config.CheckboxColumn("Email 1 Enviado"),
        "Email2_Enviado": st.column_config.CheckboxColumn("Email 2 Enviado"),
        "Email3_Enviado": st.column_config.CheckboxColumn("Email 3 Enviado"),
    }
    if sou_admin:
        col_config["Gestor_Nome"] = st.column_config.TextColumn("Gestor (nome)")
        col_config["Gestor_Email"] = st.column_config.TextColumn("Gestor (email, vai em CC)")
    else:
        # Um gestor não edita a atribuição de gestor (é sempre ele próprio) — mostra só como referência.
        col_config["Gestor_Nome"] = st.column_config.TextColumn("Gestor (nome)", disabled=True)
        col_config["Gestor_Email"] = st.column_config.TextColumn("Gestor (email)", disabled=True)

    edited = st.data_editor(
        st.session_state.clientes,
        num_rows="dynamic",
        use_container_width=True,
        column_config=col_config,
        key="editor_clientes",
    )
    if st.button("💾 Guardar alterações à tabela"):
        persistir_clientes(edited)
        st.success("Tabela atualizada e guardada — os dados ficam gravados mesmo depois de fechares o browser.")
        st.rerun()

# --- Cálculo (usado em vários separadores) ------------------------------
df_calc = calcular_ppc(clean_df(st.session_state.clientes), st.session_state.params)

# --- TAB DASHBOARD ---------------------------------------------------------
with tab_dash:
    st.subheader("Visão Geral")
    if df_calc.empty:
        st.info("Ainda não há clientes. Importa ou adiciona na aba 'Clientes'.")
    else:
        elegiveis_dash = df_calc[~df_calc["Dispensado"]]
        c1, c2, c3 = st.columns(3)
        c1.metric("Total de Clientes", len(df_calc))
        c2.metric("Dispensados", int(df_calc["Dispensado"].sum()))
        c3.metric("Elegíveis para Pagamento", len(elegiveis_dash))

        st.divider()
        st.markdown("### Estado por Pagamento")
        cols = st.columns(3)
        resumo = []
        for i, n in enumerate([1, 2, 3]):
            total = len(elegiveis_dash)
            com_guia = int(sum((row["NIF"], n) in st.session_state.guias for _, row in elegiveis_dash.iterrows()))
            enviados = int(elegiveis_dash[f"Email{n}_Enviado"].sum()) if total else 0
            pendentes = total - enviados
            resumo.append({"Pagamento": f"{n}.º", "Total Elegíveis": total, "Guia Anexada (sessão atual)": com_guia,
                            "Emails Enviados": enviados, "Pendentes": pendentes})
            with cols[i]:
                st.markdown(f"**{n}.º Pagamento**")
                st.metric("Enviados", enviados, delta=f"-{pendentes} pendentes" if pendentes else "Completo", delta_color="inverse" if pendentes else "off")
                st.progress(enviados / total if total else 0)

        st.divider()
        st.markdown("### Tabela Resumo")
        st.dataframe(pd.DataFrame(resumo), use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("### Clientes com Pagamentos Pendentes")
        pag_filtro = st.selectbox("Ver pendentes de:", [1, 2, 3], format_func=lambda x: f"{x}.º Pagamento", key="dash_filtro")
        pendentes_df = elegiveis_dash[~elegiveis_dash[f"Email{pag_filtro}_Enviado"]][["NIF", "Nome", "Email", f"Pag{pag_filtro}"]]
        if pendentes_df.empty:
            st.success(f"Todos os clientes elegíveis já receberam o {pag_filtro}.º pagamento. 🎉")
        else:
            st.dataframe(pendentes_df, use_container_width=True, height=300, hide_index=True)

        st.caption("💾 Os dados dos clientes ficam guardados de forma persistente no Supabase — não se perdem ao fechar o browser, exportar, ou voltar mais tarde para tratar do 2.º ou 3.º pagamento. As guias em PDF carregadas na aba 'Guias', porém, só existem durante a sessão atual — o estado 'Guia Emitida' fica guardado, mas o ficheiro em si tens de recarregar se voltares noutro dia.")

# --- TAB 2: Cálculo ------------------------------------------------------
with tab2:
    st.subheader("Resultado do Cálculo")
    if df_calc.empty:
        st.info("Ainda não há clientes. Importe ou adicione na aba 'Clientes'.")
    else:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Nº Clientes", len(df_calc))
        c2.metric("Nº Dispensados", int(df_calc["Dispensado"].sum()))
        c3.metric("Total PPC a cobrar", f"{df_calc.loc[~df_calc['Dispensado'], 'Total_PPC'].sum():,.2f} €".replace(",", "X").replace(".", ",").replace("X", "."))
        c4.metric("Valor médio / cliente", f"{df_calc.loc[~df_calc['Dispensado'], 'Total_PPC'].mean() if (~df_calc['Dispensado']).any() else 0:,.2f} €".replace(",", "X").replace(".", ",").replace("X", "."))

        def highlight_dispensado(row):
            return ["background-color: #E2EFDA" if row["Dispensado"] else "" for _ in row]

        show_cols = ["NIF", "Nome", "Volume_2025", "Coleta_2025", "Retencoes_2025",
                     "Base_Calculo", "Taxa", "Total_PPC", "Dispensado", "Pag1", "Pag2", "Pag3"]
        st.dataframe(
            df_calc[show_cols].style.apply(highlight_dispensado, axis=1).format(
                {"Volume_2025": "{:,.2f}", "Coleta_2025": "{:,.2f}", "Retencoes_2025": "{:,.2f}",
                 "Base_Calculo": "{:,.2f}", "Taxa": "{:.0%}", "Total_PPC": "{:,.2f}",
                 "Pag1": "{:,.2f}", "Pag2": "{:,.2f}", "Pag3": "{:,.2f}"}
            ),
            use_container_width=True,
            height=420,
        )
        st.caption("Validar o cálculo contra o simulador da OCC em alguns casos reais antes de confiar 100% na fórmula.")

# --- TAB 3: Guias ---------------------------------------------------------
with tab3:
    st.subheader("Associar Guias (PDF) aos Clientes")
    st.caption("Carregue os PDFs das guias — se o nome do ficheiro contiver o NIF (9 dígitos), a associação é automática. Caso contrário, associe manualmente abaixo.")

    n_pag = st.selectbox("A que pagamento correspondem estas guias?", [1, 2, 3], format_func=lambda x: f"{x}.º Pagamento")
    up_guias = st.file_uploader("Carregar guias PDF", type=["pdf"], accept_multiple_files=True, key="up_guias")

    if up_guias:
        for f in up_guias:
            nif_detetado = extrair_nif_de_filename(f.name)
            st.session_state.guias[(nif_detetado or f.name, n_pag)] = (f.name, f.read())
        st.success(f"{len(up_guias)} ficheiro(s) carregado(s).")

    # --- Associação manual --------------------------------------------
    if not df_calc.empty:
        clientes_nifs = set(df_calc["NIF"].tolist())
        # ficheiros deste pagamento cuja chave não corresponde a nenhum NIF de cliente
        # (ou seja, não foram associados automaticamente) + todos os já carregados, para permitir corrigir
        chaves_deste_pagamento = [k for k in st.session_state.guias.keys() if k[1] == n_pag]

        if chaves_deste_pagamento:
            st.markdown("**Associação manual / correção**")
            st.caption("Escolhe um ficheiro carregado e o cliente a quem pertence. Útil se o nome do PDF não tinha o NIF, ou se a associação automática ficou errada.")

            opcoes_ficheiro = {
                f"{st.session_state.guias[k][0]}"
                + (f"  (atualmente: sem cliente associado)" if k[0] not in clientes_nifs else f"  (atualmente: {k[0]})"):
                k
                for k in chaves_deste_pagamento
            }
            col_a, col_b, col_c = st.columns([2, 2, 1])
            with col_a:
                ficheiro_escolhido = st.selectbox("Ficheiro", list(opcoes_ficheiro.keys()), key="manual_ficheiro")
            with col_b:
                cliente_escolhido = st.selectbox(
                    "Cliente correto",
                    df_calc["NIF"].tolist(),
                    format_func=lambda n: f"{n} — {df_calc.loc[df_calc['NIF']==n,'Nome'].values[0]}",
                    key="manual_cliente",
                )
            with col_c:
                st.write("")
                st.write("")
                if st.button("Associar", key="btn_associar_manual"):
                    chave_antiga = opcoes_ficheiro[ficheiro_escolhido]
                    filename, filebytes = st.session_state.guias.pop(chave_antiga)
                    st.session_state.guias[(cliente_escolhido, n_pag)] = (filename, filebytes)
                    st.success(f"'{filename}' associado a {cliente_escolhido}.")
                    st.rerun()

    if not df_calc.empty:
        st.markdown("**Estado das guias por cliente:**")
        rows = []
        for _, r in df_calc.iterrows():
            tem_guia = (r["NIF"], n_pag) in st.session_state.guias
            rows.append({"NIF": r["NIF"], "Nome": r["Nome"], "Guia carregada": "✅" if tem_guia else "❌",
                         f"Guia{n_pag}_Emitida (registo manual)": r[f"Guia{n_pag}_Emitida"]})
        st.dataframe(pd.DataFrame(rows), use_container_width=True, height=300)

        if st.button(f"Marcar Guia {n_pag} como Emitida para todos os clientes com PDF carregado"):
            df_full = clean_df(st.session_state.clientes)
            for idx, r in df_full.iterrows():
                if (r["NIF"], n_pag) in st.session_state.guias:
                    df_full.at[idx, f"Guia{n_pag}_Emitida"] = True
            persistir_clientes(df_full)
            st.success("Estado atualizado e guardado.")
            st.rerun()

# --- TAB 4: Emails ---------------------------------------------------------
with tab4:
    st.subheader("Gerar e Enviar Emails")

    if df_calc.empty:
        st.info("Ainda não há clientes.")
    else:
        n_pag_email = st.selectbox("Qual pagamento?", [1, 2, 3], format_func=lambda x: f"{x}.º Pagamento", key="n_pag_email")
        tpl = st.session_state.templates[n_pag_email]

        with st.expander("✏️ Editar template deste email"):
            if sou_admin:
                tpl["assunto"] = st.text_input("Assunto", value=tpl["assunto"], key=f"assunto_{n_pag_email}")
                tpl["corpo"] = st.text_area("Corpo", value=tpl["corpo"], height=300, key=f"corpo_{n_pag_email}")
                st.caption("Placeholders disponíveis: {nome} {nif} {email} {pag1} {pag2} {pag3} {total} {data1} {data2} {data3}")
            else:
                st.caption("Os templates de email são definidos pelo administrador.")
                st.text_input("Assunto", value=tpl["assunto"], disabled=True)
                st.text_area("Corpo", value=tpl["corpo"], height=300, disabled=True)

        elegiveis = df_calc[~df_calc["Dispensado"]].copy()
        elegiveis = elegiveis[elegiveis["Email"].str.strip() != ""]

        st.markdown(f"**{len(elegiveis)} clientes elegíveis** (não dispensados, com email preenchido).")

        preview_nif = st.selectbox("Pré-visualizar cliente:", elegiveis["NIF"].tolist() if not elegiveis.empty else [])
        if preview_nif:
            row = elegiveis[elegiveis["NIF"] == preview_nif].iloc[0]
            assunto, corpo = render_template(tpl, row, st.session_state.params)
            st.text_input("Assunto (preview)", value=assunto, disabled=True)
            if row["Gestor_Email"]:
                st.caption(f"📋 CC: {row['Gestor_Nome'] or ''} <{row['Gestor_Email']}>")
            else:
                st.caption("📋 CC: — (sem gestor definido para este cliente)")
            st.text_area("Corpo (preview)", value=corpo, height=250, disabled=True)
            tem_guia = (row["NIF"], n_pag_email) in st.session_state.guias
            st.write("📎 Guia anexada:" , "✅ Sim" if tem_guia else "❌ Não carregada (aba Guias)")

        st.divider()
        st.markdown("### Configuração SMTP")
        c1, c2 = st.columns(2)
        with c1:
            smtp_host = st.text_input("Servidor SMTP", value="smtp.office365.com")
            smtp_user = st.text_input("Utilizador (email de login)")
            smtp_from = st.text_input("Remetente (From)", value=smtp_user)
        with c2:
            smtp_port = st.number_input("Porta", value=587, step=1)
            smtp_tls = st.checkbox("Usar STARTTLS (recomendado, porta 587)", value=True)
            smtp_pass = st.text_input("Password / App Password", type="password")

        st.caption("Gmail: smtp.gmail.com, porta 587, TLS — requer 'App Password'. Office365/Outlook: smtp.office365.com, porta 587, TLS. A password nunca é guardada — só usada durante o envio nesta sessão.")

        com_guia = [n for n in elegiveis["NIF"].tolist() if (n, n_pag_email) in st.session_state.guias]
        sem_guia = [n for n in elegiveis["NIF"].tolist() if n not in com_guia]
        nao_enviados = [
            n for n in elegiveis["NIF"].tolist()
            if not df_calc.loc[df_calc["NIF"] == n, f"Email{n_pag_email}_Enviado"].iloc[0]
        ]

        st.markdown(f"📎 **{len(com_guia)} de {len(elegiveis)} clientes elegíveis já têm guia anexada** para este pagamento.")
        if sem_guia:
            st.caption(f"Sem guia anexada (não vão poder ser enviados com anexo): {len(sem_guia)} cliente(s) — associa-os na aba 'Guias'.")

        multiselect_key = f"selecionados_email_{n_pag_email}"

        col_btn1, col_btn2, col_btn3 = st.columns(3)
        with col_btn1:
            if st.button("📎 Selecionar só quem tem guia anexada"):
                st.session_state[multiselect_key] = [n for n in com_guia if n in nao_enviados]
                st.rerun()
        with col_btn2:
            if st.button("☑️ Selecionar todos os elegíveis"):
                st.session_state[multiselect_key] = nao_enviados
                st.rerun()
        with col_btn3:
            if st.button("✖️ Limpar seleção"):
                st.session_state[multiselect_key] = []
                st.rerun()

        if multiselect_key not in st.session_state:
            st.session_state[multiselect_key] = [n for n in com_guia if n in nao_enviados]

        selecionados = st.multiselect(
            "Clientes selecionados para envio (podes ajustar manualmente)",
            elegiveis["NIF"].tolist(),
            format_func=lambda n: f"{n} — {elegiveis.loc[elegiveis['NIF']==n,'Nome'].values[0]}" + ("" if n in com_guia else "  ⚠️ sem guia"),
            key=multiselect_key,
        )

        if st.button("🚀 Enviar Emails Selecionados", type="primary", disabled=not selecionados):
            if not smtp_user or not smtp_pass:
                st.error("Preencher utilizador e password SMTP.")
            else:
                smtp_cfg = {
                    "host": smtp_host, "porta": int(smtp_port), "tls": smtp_tls,
                    "utilizador": smtp_user, "password": smtp_pass, "remetente": smtp_from,
                }
                progress = st.progress(0.0)
                status_box = st.empty()
                df_full = clean_df(st.session_state.clientes)
                sucessos, falhas = 0, 0
                for i, nif in enumerate(selecionados):
                    row = elegiveis[elegiveis["NIF"] == nif].iloc[0]
                    assunto, corpo = render_template(tpl, row, st.session_state.params)
                    anexos = []
                    guia = st.session_state.guias.get((nif, n_pag_email))
                    if guia:
                        anexos.append(guia)
                    try:
                        cc_gestor = [row["Gestor_Email"]] if row["Gestor_Email"] else []
                        enviar_email(smtp_cfg, row["Email"], assunto, corpo, anexos, cc=cc_gestor)
                        idx = df_full.index[df_full["NIF"] == nif][0]
                        df_full.at[idx, f"Email{n_pag_email}_Enviado"] = True
                        registar_log(
                            {"data": datetime.now().strftime("%Y-%m-%d %H:%M"), "nif": nif,
                             "nome": row["Nome"], "pagamento": n_pag_email, "estado": "Enviado"}
                        )
                        sucessos += 1
                    except Exception as e:
                        registar_log(
                            {"data": datetime.now().strftime("%Y-%m-%d %H:%M"), "nif": nif,
                             "nome": row["Nome"], "pagamento": n_pag_email, "estado": f"Erro: {e}"}
                        )
                        falhas += 1
                    progress.progress((i + 1) / len(selecionados))
                    status_box.text(f"{i+1}/{len(selecionados)} — {row['Nome']}")
                persistir_clientes(df_full)
                st.success(f"Concluído: {sucessos} enviados, {falhas} com erro. Estados guardados.")
                st.rerun()

        if st.session_state.log_envio:
            st.markdown("### Log de Envios")
            st.dataframe(pd.DataFrame(st.session_state.log_envio), use_container_width=True, height=250)

# --- TAB 5: Exportar --------------------------------------------------------
with tab5:
    st.subheader("Exportar Folha de Controlo")
    st.caption("📌 Exportar é apenas um download — os teus dados continuam guardados na app depois disto, para poderes continuar a usá-la ao longo do ano para o 2.º e 3.º pagamento.")
    if df_calc.empty:
        st.info("Ainda não há clientes.")
    else:
        excel_bytes = gerar_excel(df_calc, st.session_state.params)
        st.download_button(
            "⬇️ Descarregar Excel de Controlo (com fórmulas e estados)",
            data=excel_bytes,
            file_name=f"Controlo_PPC_2026_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.caption("O Excel inclui todos os cálculos, os estados de Guia Emitida / Email Enviado e fica destacado a verde para clientes dispensados.")

        if st.session_state.log_envio:
            log_csv = pd.DataFrame(st.session_state.log_envio).to_csv(index=False, sep=";")
            st.download_button("⬇️ Descarregar log de envios (CSV)", log_csv, file_name="log_envios_ppc.csv", mime="text/csv")

# --- TAB GESTORES (só admin) ------------------------------------------------
if sou_admin and tab_gestores is not None:
    with tab_gestores:
        st.subheader("Contas de Gestor")
        st.caption("Cria e gere as contas de login dos gestores da SERVE. Depois de criares a conta, atribui os clientes a este gestor na aba 'Clientes' (campo Gestor_Email igual ao email de login abaixo).")

        client = get_client()
        perfis_resp = client.table("perfis").select("*").order("email").execute()
        perfis_lista = perfis_resp.data or []

        if perfis_lista:
            st.dataframe(
                pd.DataFrame(perfis_lista)[["email", "nome", "role"]].rename(
                    columns={"email": "Email", "nome": "Nome", "role": "Papel"}
                ),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.info("Ainda não há contas registadas além da tua.")

        st.divider()
        st.markdown("### Criar nova conta de gestor")

        if not SUPABASE_SERVICE_KEY:
            st.warning(
                "Falta configurar SUPABASE_SERVICE_KEY em Settings → Secrets para poderes criar contas "
                "diretamente pela app. Ver GUIA_SUPABASE.md, secção 'Adicionar gestores'."
            )
        else:
            with st.form("form_novo_gestor"):
                novo_nome = st.text_input("Nome do gestor")
                novo_email = st.text_input("Email de login")
                nova_pass = st.text_input("Password inicial (o gestor pode alterá-la depois)", type="password")
                novo_role = st.selectbox("Papel", ["gestor", "admin"])
                submitted = st.form_submit_button("Criar conta")
                if submitted:
                    if not novo_email or not nova_pass:
                        st.error("Preenche pelo menos o email e a password.")
                    else:
                        try:
                            admin_client = get_admin_client()
                            criado = admin_client.auth.admin.create_user(
                                {
                                    "email": novo_email,
                                    "password": nova_pass,
                                    "email_confirm": True,
                                    "user_metadata": {"nome": novo_nome},
                                }
                            )
                            admin_client.table("perfis").upsert(
                                {"id": criado.user.id, "email": novo_email, "nome": novo_nome, "role": novo_role}
                            ).execute()
                            st.success(f"Conta criada para {novo_email}. Já pode entrar na app com esta password.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Não foi possível criar a conta: {e}")

# ---------------------------------------------------------------------------
# Persistir parâmetros e templates (só o admin escreve; RLS bloqueia o resto)
# ---------------------------------------------------------------------------
guardar_config_db(st.session_state.params, st.session_state.templates)

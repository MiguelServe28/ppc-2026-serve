"""
Lógica partilhada por toda a plataforma SERVE (registo central de clientes,
ligação ao Supabase, autenticação, e utilitários usados por mais do que um
imposto). Cada página em paginas/ importa deste módulo.

Não corras este ficheiro diretamente — é só uma biblioteca. O ponto de entrada
da app é o app.py.
"""

import io
import math
import re
import smtplib
import ssl
from datetime import date
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pandas as pd
import pdfplumber
import streamlit as st
from cryptography.fernet import Fernet, InvalidToken
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from supabase import Client, create_client

# ---------------------------------------------------------------------------
# Registo central de clientes (partilhado por toda a plataforma)
# ---------------------------------------------------------------------------
CLIENT_COLS = [
    "NIF", "Numero_Cliente", "Nome", "Email", "Lingua", "Gestor_Nome", "Gestor_Email",
    "Tipo_Empresa", "Tipo_AL", "Tipo_Trab_Independente", "Tipo_Rep_Fiscal",
    "Aplica_PPC", "Aplica_IVA", "Aplica_IMI", "Aplica_IRS", "Aplica_SS",
    "IRS_Avulso", "Notas",
]
TIPO_COLS = ["Tipo_Empresa", "Tipo_AL", "Tipo_Trab_Independente", "Tipo_Rep_Fiscal"]
APLICA_COLS = ["Aplica_PPC", "Aplica_IVA", "Aplica_IMI", "Aplica_IRS", "Aplica_SS"]
# IRS_Avulso: cliente importado apenas pelo menu do IRS (não é cliente de avença
# da base central) — permite separar "avença" de "só IRS" na Visão Geral do IRS.
BOOL_COLS = TIPO_COLS + APLICA_COLS + ["IRS_Avulso"]
# Lingua: "PT" ou "EN" — decide em que língua os emails deste cliente são enviados.
TEXT_COLS = ["NIF", "Numero_Cliente", "Nome", "Email", "Lingua", "Gestor_Nome", "Gestor_Email", "Notas"]

COLUMN_MAP_TO_DB = {
    "NIF": "nif", "Numero_Cliente": "numero_cliente", "Nome": "nome", "Email": "email",
    "Lingua": "lingua", "Gestor_Nome": "gestor_nome", "Gestor_Email": "gestor_email",
    "Tipo_Empresa": "tipo_empresa", "Tipo_AL": "tipo_al",
    "Tipo_Trab_Independente": "tipo_trabalhador_independente", "Tipo_Rep_Fiscal": "tipo_representacao_fiscal",
    "Aplica_PPC": "aplica_ppc", "Aplica_IVA": "aplica_iva", "Aplica_IMI": "aplica_imi",
    "Aplica_IRS": "aplica_irs", "Aplica_SS": "aplica_ss",
    "IRS_Avulso": "irs_avulso", "Notas": "notas",
}
COLUMN_MAP_FROM_DB = {v: k for k, v in COLUMN_MAP_TO_DB.items()}

# ---------------------------------------------------------------------------
# Dados específicos do PPC (tabela própria, ligada por NIF)
# ---------------------------------------------------------------------------
PPC_COLS = [
    "NIF", "Volume", "Coleta", "Retencoes",
    "Guia1_Emitida", "Guia2_Emitida", "Guia3_Emitida",
    "Email1_Enviado", "Email2_Enviado", "Email3_Enviado",
]
PPC_BOOL_COLS = [c for c in PPC_COLS if c.startswith("Guia") or c.startswith("Email")]
PPC_NUM_COLS = ["Volume", "Coleta", "Retencoes"]

PPC_COLUMN_MAP_TO_DB = {
    "NIF": "nif", "Volume": "volume", "Coleta": "coleta", "Retencoes": "retencoes",
    "Guia1_Emitida": "guia1_emitida", "Guia2_Emitida": "guia2_emitida", "Guia3_Emitida": "guia3_emitida",
    "Email1_Enviado": "email1_enviado", "Email2_Enviado": "email2_enviado", "Email3_Enviado": "email3_enviado",
}
PPC_COLUMN_MAP_FROM_DB = {v: k for k, v in PPC_COLUMN_MAP_TO_DB.items()}

DEFAULT_TEMPLATES = {
    1: {
        "assunto": "Pagamentos por Conta {ano_pagamentos} — {nome}",
        "corpo": (
            "Exmo(a). Sr(a).,\n\n"
            "No seguimento do apuramento do IRC referente a {ano_dados}, informamos que a {nome} "
            "(NIF {nif}) tem pagamentos por conta a efetuar em {ano_pagamentos}, nos seguintes montantes e prazos:\n\n"
            "• 1.º Pagamento por Conta: {pag1} € — até {data1}\n"
            "• 2.º Pagamento por Conta: {pag2} € — até {data2}\n"
            "• 3.º Pagamento por Conta: {pag3} € — até {data3}\n\n"
            "Total anual: {total} €\n\n"
            "Segue em anexo a guia referente ao 1.º pagamento. Solicitamos que proceda ao pagamento até à data "
            "indicada, de forma a evitar juros de mora.\n\n"
            "As guias do 2.º e 3.º pagamento serão enviadas atempadamente.\n\n"
            "Ficamos ao dispor para qualquer esclarecimento.\n\n"
            "Com os melhores cumprimentos,"
        ),
        "assunto_en": "Payments on Account {ano_pagamentos} — {nome}",
        "corpo_en": (
            "Dear Sir or Madam,\n\n"
            "Following the assessment of the {ano_dados} Corporate Income Tax (IRC), we inform you that {nome} "
            "(tax no. {nif}) has payments on account due in {ano_pagamentos}, in the following amounts and deadlines:\n\n"
            "• 1st Payment on Account: {pag1} € — by {data1}\n"
            "• 2nd Payment on Account: {pag2} € — by {data2}\n"
            "• 3rd Payment on Account: {pag3} € — by {data3}\n\n"
            "Annual total: {total} €\n\n"
            "Please find attached the payment form for the 1st payment. We kindly ask you to settle it by the date "
            "indicated in order to avoid late-payment interest.\n\n"
            "The payment forms for the 2nd and 3rd payments will be sent in due course.\n\n"
            "We remain at your disposal for any clarification.\n\n"
            "Best regards,"
        ),
    },
    2: {
        "assunto": "2.º Pagamento por Conta {ano_pagamentos} — {nome}",
        "corpo": (
            "Exmo(a). Sr(a).,\n\n"
            "No seguimento do 1.º pagamento por conta já efetuado, relembramos que o 2.º pagamento por conta "
            "da {nome} (NIF {nif}) vence a {data2}, no valor de {pag2} €.\n\n"
            "Segue em anexo a respetiva guia.\n\n"
            "Ficamos ao dispor para qualquer esclarecimento.\n\n"
            "Com os melhores cumprimentos,"
        ),
        "assunto_en": "2nd Payment on Account {ano_pagamentos} — {nome}",
        "corpo_en": (
            "Dear Sir or Madam,\n\n"
            "Following the 1st payment on account already made, we remind you that the 2nd payment on account "
            "of {nome} (tax no. {nif}) is due on {data2}, in the amount of {pag2} €.\n\n"
            "Please find the payment form attached.\n\n"
            "We remain at your disposal for any clarification.\n\n"
            "Best regards,"
        ),
    },
    3: {
        "assunto": "3.º Pagamento por Conta {ano_pagamentos} — {nome}",
        "corpo": (
            "Exmo(a). Sr(a).,\n\n"
            "No seguimento dos pagamentos por conta já efetuados, relembramos que o 3.º e último pagamento por "
            "conta da {nome} (NIF {nif}) vence a {data3}, no valor de {pag3} €.\n\n"
            "Segue em anexo a respetiva guia.\n\n"
            "Ficamos ao dispor para qualquer esclarecimento.\n\n"
            "Com os melhores cumprimentos,"
        ),
        "assunto_en": "3rd Payment on Account {ano_pagamentos} — {nome}",
        "corpo_en": (
            "Dear Sir or Madam,\n\n"
            "Following the payments on account already made, we remind you that the 3rd and final payment on "
            "account of {nome} (tax no. {nif}) is due on {data3}, in the amount of {pag3} €.\n\n"
            "Please find the payment form attached.\n\n"
            "We remain at your disposal for any clarification.\n\n"
            "Best regards,"
        ),
    },
}

# ---------------------------------------------------------------------------
# Ligação ao Supabase
# ---------------------------------------------------------------------------
SUPABASE_URL = st.secrets.get("SUPABASE_URL")
SUPABASE_ANON_KEY = st.secrets.get("SUPABASE_ANON_KEY")
SUPABASE_SERVICE_KEY = st.secrets.get("SUPABASE_SERVICE_KEY")  # só necessário para o admin criar gestores


def get_client() -> Client:
    """Devolve o cliente Supabase desta sessão de browser (um por utilizador —
    nunca partilhado entre utilizadores, para que a sessão de autenticação de
    cada um não se misture com a de outro)."""
    if "sb_client" not in st.session_state:
        st.session_state.sb_client = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
    return st.session_state.sb_client


def get_admin_client() -> "Client | None":
    """Cliente com a chave 'service_role' — só usado nas ações de administração
    de contas (criar gestores). Nunca deve ser usado para ler/escrever clientes."""
    if not SUPABASE_SERVICE_KEY:
        return None
    return create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)


# ---------------------------------------------------------------------------
# Autenticação e perfil (admin vs. gestor)
# ---------------------------------------------------------------------------
def carregar_perfil():
    client = get_client()
    user_id = st.session_state.user.id
    resp = client.table("perfis").select("*").eq("id", user_id).execute()
    if resp.data:
        st.session_state.perfil = resp.data[0]
    else:
        st.session_state.perfil = {"id": user_id, "email": st.session_state.user.email, "nome": st.session_state.user.email, "role": "gestor"}


def sou_admin() -> bool:
    return "perfil" in st.session_state and st.session_state.perfil["role"] == "admin"


def meu_email() -> str:
    return st.session_state.perfil["email"]


def ecra_login():
    st.title("🔒 Gestão Fiscal SERVE")
    st.caption("SERVE — Contabilidade e Viabilização Empresarial. Inicia sessão com a tua conta de gestor.")

    if not SUPABASE_URL or not SUPABASE_ANON_KEY:
        st.error(
            "A app ainda não está ligada ao Supabase. Falta configurar SUPABASE_URL e "
            "SUPABASE_ANON_KEY em Settings → Secrets. Ver GUIA_SUPABASE.md."
        )
        st.stop()

    email = st.text_input("Email")
    pwd = st.text_input("Password", type="password")
    if st.button("Entrar", type="primary"):
        try:
            client = get_client()
            res = client.auth.sign_in_with_password({"email": email, "password": pwd})
            st.session_state.user = res.user
            carregar_perfil()
            st.rerun()
        except Exception:
            st.error("Email ou password incorretos.")


def sidebar_utilizador():
    perfil = st.session_state.perfil
    with st.sidebar:
        papel = "Administrador" if perfil["role"] == "admin" else "Gestor"
        st.success(f"👤 {perfil['nome'] or perfil['email']}  \n**{papel}**")
        if st.button("Sair"):
            try:
                get_client().auth.sign_out()
            except Exception:
                pass
            for k in ("user", "perfil", "sb_client", "clientes", "ppc_dados", "irs_dados",
                      "params", "templates", "template_irs", "log_envio", "guias_por_associar"):
                st.session_state.pop(k, None)
            st.rerun()
        st.divider()


def requer_login():
    """Chama-se no topo de cada página. Se não houver sessão, mostra o login e para."""
    if "user" not in st.session_state:
        ecra_login()
        st.stop()
    sidebar_utilizador()


# ---------------------------------------------------------------------------
# Persistência — registo central de clientes (Supabase)
# ---------------------------------------------------------------------------
def clean_clientes_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in CLIENT_COLS:
        if c not in df.columns:
            df[c] = False if c in BOOL_COLS else ""
    for c in BOOL_COLS:
        df[c] = df[c].fillna(False).astype(bool)
    for c in TEXT_COLS:
        df[c] = df[c].fillna("").astype(str).str.strip()
    # Normaliza a língua: aceita "pt", "en", "Português", "English", etc. — tudo
    # o que não começar por EN fica PT (a língua por omissão).
    df["Lingua"] = df["Lingua"].str.upper().str[:2]
    df.loc[~df["Lingua"].isin(["PT", "EN"]), "Lingua"] = "PT"
    return df[CLIENT_COLS]


def clean_ppc_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in PPC_COLS:
        if c not in df.columns:
            df[c] = False if c in PPC_BOOL_COLS else ("" if c == "NIF" else 0.0)
    df["NIF"] = df["NIF"].fillna("").astype(str).str.strip()
    for c in PPC_NUM_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    for c in PPC_BOOL_COLS:
        df[c] = df[c].fillna(False).astype(bool)
    df = df[df["NIF"] != ""]
    return df[PPC_COLS]


def carregar_clientes_db() -> pd.DataFrame:
    client = get_client()
    resp = client.table("clientes").select("*").execute()
    rows = resp.data or []
    if not rows:
        return pd.DataFrame(columns=CLIENT_COLS)
    df = pd.DataFrame(rows).rename(columns=COLUMN_MAP_FROM_DB)
    return clean_clientes_df(df)


def perfil_nome_vazio() -> bool:
    return not (st.session_state.perfil.get("nome") or "").strip()


def guardar_clientes_db(df: pd.DataFrame, nifs_antes: set = None):
    """Grava por DIFERENÇA (nunca 'apaga tudo primeiro'): faz upsert das linhas
    de df e apaga apenas os NIFs que existiam antes e desapareceram. Assim, se a
    ligação falhar a meio, o pior que acontece é a gravação ficar incompleta —
    nunca se perde a carteira inteira. Um gestor só toca nos seus próprios
    clientes (RLS trata do âmbito)."""
    client = get_client()
    df2 = clean_clientes_df(df).copy()

    if not sou_admin():
        df2["Gestor_Email"] = meu_email()
        if not perfil_nome_vazio():
            df2["Gestor_Nome"] = st.session_state.perfil["nome"]

    if nifs_antes is None:
        antes = clean_clientes_df(st.session_state.get("clientes", pd.DataFrame(columns=CLIENT_COLS)))
        nifs_antes = set(antes["NIF"])
    para_apagar = nifs_antes - set(df2["NIF"])
    if para_apagar:
        client.table("clientes").delete().in_("nif", list(para_apagar)).execute()
    if not df2.empty:
        registos = df2.rename(columns=COLUMN_MAP_TO_DB).to_dict("records")
        client.table("clientes").upsert(registos, on_conflict="nif").execute()


def persistir_clientes(df: pd.DataFrame):
    """Grava no Supabase E atualiza a sessão. Usa-se quando 'df' é o conjunto
    COMPLETO de clientes pretendido (ex: importação) — substitui tudo, mas de
    forma segura (por diferença)."""
    df = clean_clientes_df(df)
    if not sou_admin():
        df["Gestor_Email"] = meu_email()
        if not perfil_nome_vazio():
            df["Gestor_Nome"] = st.session_state.perfil["nome"]
    antes = clean_clientes_df(st.session_state.get("clientes", pd.DataFrame(columns=CLIENT_COLS)))
    guardar_clientes_db(df, nifs_antes=set(antes["NIF"]))
    st.session_state.clientes = df


def guardar_clientes_parcial_db(df_editado: pd.DataFrame, nifs_visiveis_antes: set) -> pd.DataFrame:
    """Grava só as alterações dentro de um subconjunto filtrado (ex: 'Só IRS'),
    sem tocar em clientes fora desse filtro. Compara os NIFs que estavam
    visíveis antes da edição com os que sobram depois — os que desapareceram
    do filtro são apagados, os restantes são upsert. Devolve o df já limpo."""
    client = get_client()
    df2 = clean_clientes_df(df_editado).copy()
    if not sou_admin():
        df2["Gestor_Email"] = meu_email()
        if not perfil_nome_vazio():
            df2["Gestor_Nome"] = st.session_state.perfil["nome"]

    nifs_finais = set(df2["NIF"])
    nifs_para_apagar = nifs_visiveis_antes - nifs_finais
    if nifs_para_apagar:
        client.table("clientes").delete().in_("nif", list(nifs_para_apagar)).execute()
    if not df2.empty:
        registos = df2.rename(columns=COLUMN_MAP_TO_DB).to_dict("records")
        client.table("clientes").upsert(registos, on_conflict="nif").execute()
    return df2


def persistir_clientes_parcial(df_editado: pd.DataFrame, nifs_visiveis_antes: set):
    """Versão segura para usar com uma tabela FILTRADA (ex: só clientes de IRS):
    só toca nas linhas que estavam visíveis nesse filtro, nunca nas restantes."""
    df2 = guardar_clientes_parcial_db(df_editado, nifs_visiveis_antes)
    completo = clean_clientes_df(st.session_state.clientes)
    resto = completo[~completo["NIF"].isin(nifs_visiveis_antes)]
    st.session_state.clientes = clean_clientes_df(pd.concat([resto, df2], ignore_index=True))


# ---------------------------------------------------------------------------
# Persistência — dados específicos do PPC
# ---------------------------------------------------------------------------
def carregar_ppc_db() -> pd.DataFrame:
    client = get_client()
    resp = client.table("ppc_dados").select("*").execute()
    rows = resp.data or []
    if not rows:
        return pd.DataFrame(columns=PPC_COLS)
    df = pd.DataFrame(rows).rename(columns=PPC_COLUMN_MAP_FROM_DB)
    return clean_ppc_df(df)


def guardar_ppc_db(df: pd.DataFrame, nifs_antes: set = None):
    """Grava por diferença (upsert + apagar só os que desapareceram) — ver nota
    em guardar_clientes_db."""
    client = get_client()
    df2 = clean_ppc_df(df).copy()
    if nifs_antes is None:
        antes = clean_ppc_df(st.session_state.get("ppc_dados", pd.DataFrame(columns=PPC_COLS)))
        nifs_antes = set(antes["NIF"])
    para_apagar = nifs_antes - set(df2["NIF"])
    if para_apagar:
        client.table("ppc_dados").delete().in_("nif", list(para_apagar)).execute()
    if not df2.empty:
        registos = df2.rename(columns=PPC_COLUMN_MAP_TO_DB).to_dict("records")
        client.table("ppc_dados").upsert(registos, on_conflict="nif").execute()


def persistir_ppc(df: pd.DataFrame):
    df = clean_ppc_df(df)
    guardar_ppc_db(df)
    st.session_state.ppc_dados = df


def montar_base_ppc() -> pd.DataFrame:
    """Junta os clientes com 'Aplica_PPC' ligado com os respetivos dados de PPC
    (mesmo que ainda não tenham nenhum valor preenchido)."""
    clientes = clean_clientes_df(st.session_state.clientes)
    elegiveis = clientes[clientes["Aplica_PPC"]].copy()
    ppc = clean_ppc_df(st.session_state.ppc_dados)
    base = elegiveis.merge(ppc, on="NIF", how="left")
    for c in PPC_NUM_COLS:
        base[c] = pd.to_numeric(base[c], errors="coerce").fillna(0.0)
    for c in PPC_BOOL_COLS:
        base[c] = base[c].fillna(False).astype(bool)
    return base


# ---------------------------------------------------------------------------
# Persistência — log de envios e configuração
# ---------------------------------------------------------------------------
def carregar_log_db() -> list:
    client = get_client()
    resp = client.table("log_envios").select("data, nif, nome, pagamento, estado, modulo, enviado_por").order("id").execute()
    return resp.data or []


def guardar_log_entry_db(entry: dict):
    client = get_client()
    client.table("log_envios").insert(entry).execute()


def registar_log(entry: dict):
    st.session_state.log_envio.append(entry)
    guardar_log_entry_db(entry)


def carregar_config_db():
    client = get_client()
    try:
        resp = client.table("config").select("params_json, templates_json, templates_irs_json, templates_ss_json").eq("id", 1).execute()
    except Exception:
        # A coluna templates_ss_json só existe a partir do v7 — se ainda não
        # foi corrido, carrega sem ela para a app não deixar de funcionar.
        resp = client.table("config").select("params_json, templates_json, templates_irs_json").eq("id", 1).execute()
    if not resp.data:
        return None, None, None, None
    row = resp.data[0]
    params_json, templates_json = row.get("params_json"), row.get("templates_json")
    templates_irs_json = row.get("templates_irs_json")
    template_ss = row.get("templates_ss_json")
    params, templates, template_irs = None, None, None
    if params_json:
        params = {
            "limiar_volume": params_json.get("limiar_volume", 500000.0), "taxa_baixa": params_json.get("taxa_baixa", 0.80),
            "taxa_alta": params_json.get("taxa_alta", 0.95), "limite_dispensa": params_json.get("limite_dispensa", 200.0),
            "ano_dados": int(params_json.get("ano_dados", 2025)),
            "ano_pagamentos": int(params_json.get("ano_pagamentos", 2026)),
            "assinatura_html": params_json.get("assinatura_html", ""),
            "data1": date.fromisoformat(params_json["data1"]), "data2": date.fromisoformat(params_json["data2"]),
            "data3": date.fromisoformat(params_json["data3"]),
        }
    if templates_json:
        templates = {int(k): v for k, v in templates_json.items()}
    if templates_irs_json:
        template_irs = templates_irs_json
    return params, templates, template_irs, template_ss


def guardar_config_db(params: dict, templates: dict, template_irs: dict = None, template_ss: dict = None):
    """Só o admin tem permissão (RLS) para escrever a configuração global."""
    if not sou_admin():
        return
    params_serializ = dict(params)
    for k in ("data1", "data2", "data3"):
        params_serializ[k] = params[k].isoformat()
    templates_serializ = {str(k): v for k, v in templates.items()}
    registo = {"id": 1, "params_json": params_serializ, "templates_json": templates_serializ}
    if template_irs is not None:
        registo["templates_irs_json"] = template_irs
    if template_ss is not None:
        registo["templates_ss_json"] = template_ss
    client = get_client()
    try:
        client.table("config").upsert(registo).execute()
    except Exception:
        # Sem a coluna do v7 ainda, grava o resto na mesma.
        registo.pop("templates_ss_json", None)
        client.table("config").upsert(registo).execute()


# ---------------------------------------------------------------------------
# Estado (carregado uma vez por sessão)
# ---------------------------------------------------------------------------
def init_state():
    if "perfil" not in st.session_state:
        carregar_perfil()
    if "clientes" not in st.session_state:
        st.session_state.clientes = carregar_clientes_db()
    if "ppc_dados" not in st.session_state:
        st.session_state.ppc_dados = carregar_ppc_db()
    if "irs_dados" not in st.session_state:
        st.session_state.irs_dados = carregar_irs_db()
    if "guias_por_associar" not in st.session_state:
        # PDFs de guias PPC carregados cujo nome não tinha NIF — ficam aqui à
        # espera de associação manual; depois de associados vão para o Storage.
        st.session_state.guias_por_associar = {}  # {(n_pag, filename): bytes}
    if ("params" not in st.session_state or "templates" not in st.session_state
            or "template_irs" not in st.session_state or "template_ss" not in st.session_state):
        params_db, templates_db, template_irs_db, template_ss_db = carregar_config_db()
        if "params" not in st.session_state:
            st.session_state.params = params_db or {
                "limiar_volume": 500000.0,
                "taxa_baixa": 0.80,
                "taxa_alta": 0.95,
                "limite_dispensa": 200.0,
                "ano_dados": 2025,        # ano dos dados (Modelo 22 / liquidações)
                "ano_pagamentos": 2026,   # ano em que os pagamentos são feitos
                "assinatura_html": "",    # assinatura acrescentada aos emails (HTML simples)
                "data1": date(2026, 7, 31),
                "data2": date(2026, 9, 30),
                "data3": date(2026, 12, 15),
            }
        if "templates" not in st.session_state:
            st.session_state.templates = templates_db or {k: v.copy() for k, v in DEFAULT_TEMPLATES.items()}
        if "template_irs" not in st.session_state:
            st.session_state.template_irs = template_irs_db or DEFAULT_TEMPLATE_IRS.copy()
        if "template_ss" not in st.session_state:
            st.session_state.template_ss = template_ss_db or DEFAULT_TEMPLATE_SS.copy()
    if "log_envio" not in st.session_state:
        st.session_state.log_envio = carregar_log_db()


# ---------------------------------------------------------------------------
# Importação de ficheiros (CSV/Excel)
# ---------------------------------------------------------------------------
def parse_numero_pt(serie: pd.Series) -> pd.Series:
    """Converte números em formato português (vírgula decimal, ponto de milhares) ou internacional para float."""
    def conv(v):
        if pd.isna(v):
            return None
        s = str(v).strip()
        if s == "":
            return None
        s = s.replace(" ", "").replace("€", "")
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return serie.apply(conv)


def ler_ficheiro_importacao(uploaded_file) -> pd.DataFrame:
    """Lê CSV ou Excel de forma tolerante: deteta o encoding do CSV automaticamente
    (ficheiros exportados do Excel em português costumam vir em Windows-1252/ISO-8859-1,
    não UTF-8) e trata números com vírgula decimal."""
    nome = uploaded_file.name.lower()
    if nome.endswith(".csv"):
        raw = uploaded_file.read()
        texto = None
        for enc in ("utf-8-sig", "utf-8", "cp1252", "iso-8859-1"):
            try:
                texto = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if texto is None:
            texto = raw.decode("latin1", errors="replace")
        df = pd.read_csv(io.StringIO(texto), sep=None, engine="python")
    else:
        df = pd.read_excel(uploaded_file)

    df.columns = [str(c).strip() for c in df.columns]
    # Aceita tanto os nomes novos (Volume, Coleta, Retencoes) como os antigos
    # com ano (Volume_2025, ...) ou com o ano configurado (Volume_2026, ...).
    aliases = {}
    for base in ("Volume", "Coleta", "Retencoes"):
        for c in df.columns:
            if c == base or re.fullmatch(rf"{base}_\d{{4}}", c):
                aliases[c] = base
    # O N.º interno do cliente pode vir escrito de várias formas.
    for c in df.columns:
        if c in ("N.º", "Nº", "N.o", "N°", "No.", "Numero", "Número", "Num", "Numero_Cliente", "N.º Cliente"):
            aliases[c] = "Numero_Cliente"
    df = df.rename(columns=aliases)
    if "Numero_Cliente" in df.columns:
        df["Numero_Cliente"] = (
            df["Numero_Cliente"].fillna("").astype(str).str.strip().str.replace(r"\.0$", "", regex=True)
        )
        df.loc[df["Numero_Cliente"] == "nan", "Numero_Cliente"] = ""
    for c in ("Volume", "Coleta", "Retencoes"):
        if c in df.columns:
            df[c] = parse_numero_pt(df[c])
    if "Email" in df.columns:
        df["Email"] = df["Email"].astype(str).str.strip().str.rstrip(",").str.strip()
        df.loc[df["Email"] == "nan", "Email"] = ""
    if "NIF" in df.columns:
        df["NIF"] = df["NIF"].astype(str).str.strip()
    return df


# ---------------------------------------------------------------------------
# Cálculo PPC
# ---------------------------------------------------------------------------
def calcular_ppc(df: pd.DataFrame, params: dict) -> pd.DataFrame:
    df = df.copy()
    df["Base_Calculo"] = (df["Coleta"] - df["Retencoes"]).clip(lower=0)
    df["Taxa"] = df["Volume"].apply(
        lambda v: params["taxa_baixa"] if v <= params["limiar_volume"] else params["taxa_alta"]
    )
    df["Dispensado"] = df["Base_Calculo"] < params["limite_dispensa"]
    df["Total_PPC_Base"] = df["Base_Calculo"] * df["Taxa"]

    def parcela(row):
        if row["Dispensado"] or row["Total_PPC_Base"] <= 0:
            return 0
        return math.ceil(round(row["Total_PPC_Base"] / 3, 6) - 1e-9)

    df["Pag1"] = df.apply(parcela, axis=1)
    df["Pag2"] = df["Pag1"]
    df["Pag3"] = df["Pag1"]
    df["Total_PPC"] = df["Pag1"] + df["Pag2"] + df["Pag3"]
    return df


# ---------------------------------------------------------------------------
# Export Excel (folha de controlo do PPC)
# ---------------------------------------------------------------------------
def gerar_excel_ppc(df_calc: pd.DataFrame, params: dict) -> bytes:
    ano_dados = params.get("ano_dados", 2025)
    ano_pag = params.get("ano_pagamentos", 2026)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Controlo PPC {ano_pag}"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    FONT = "Arial"
    HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    DISPENSA_FILL = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    HEADER_FONT = Font(name=FONT, color="FFFFFF", bold=True, size=10)
    TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
    BLACK = Font(name=FONT, color="000000")
    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"Controlo de Pagamentos por Conta — {ano_pag}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    headers = [
        "NIF", "Nome", "Email", "Gestor (nome)", "Gestor (email)", f"Volume {ano_dados}", f"Coleta {ano_dados}", f"Retenções {ano_dados}",
        "Base de Cálculo", "Taxa", "Total PPC", "Dispensado",
        "1º Pagamento", "2º Pagamento", "3º Pagamento",
        "Data Limite 1º", "Data Limite 2º", "Data Limite 3º",
        "Guia1 Emitida", "Guia2 Emitida", "Guia3 Emitida",
        "Email1 Enviado", "Email2 Enviado", "Email3 Enviado", "Notas",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[4].height = 30

    widths = [12, 26, 24, 18, 24, 13, 13, 13, 13, 8, 12, 11, 11, 11, 11, 12, 12, 12, 11, 11, 11, 11, 11, 11, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 5
    for _, r in df_calc.iterrows():
        vals = [
            r["NIF"], r["Nome"], r["Email"], r["Gestor_Nome"], r["Gestor_Email"],
            r["Volume"], r["Coleta"], r["Retencoes"],
            r["Base_Calculo"], r["Taxa"], r["Total_PPC"], "Sim" if r["Dispensado"] else "Não",
            r["Pag1"], r["Pag2"], r["Pag3"],
            params["data1"] if not r["Dispensado"] else None,
            params["data2"] if not r["Dispensado"] else None,
            params["data3"] if not r["Dispensado"] else None,
            "Sim" if r["Guia1_Emitida"] else "Não",
            "Sim" if r["Guia2_Emitida"] else "Não",
            "Sim" if r["Guia3_Emitida"] else "Não",
            "Sim" if r["Email1_Enviado"] else "Não",
            "Sim" if r["Email2_Enviado"] else "Não",
            "Sim" if r["Email3_Enviado"] else "Não",
            r["Notas"],
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = BLACK
            c.border = BORDER
            if i in (6, 7, 8, 9, 11, 13, 14, 15):
                c.number_format = "#,##0.00"
            if i == 10:
                c.number_format = "0%"
            if i in (16, 17, 18):
                c.number_format = "dd/mm/yyyy"
        if r["Dispensado"]:
            for i in range(1, len(vals) + 1):
                ws.cell(row=row, column=i).fill = DISPENSA_FILL
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Emails
# ---------------------------------------------------------------------------
def formatar_valor(v):
    return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def lingua_cliente(row) -> str:
    """Língua do cliente ('PT' ou 'EN') — vem do registo central. Tudo o que
    não for EN conta como PT."""
    try:
        valor = str(row.get("Lingua", "PT") or "PT")
    except Exception:
        valor = "PT"
    return "EN" if valor.strip().upper().startswith("EN") else "PT"


def texto_template(template: dict, chave: str, lingua: str) -> str:
    """Devolve o texto do template na língua pedida. Se a versão EN estiver
    vazia (ainda não preenchida), usa a PT — nunca falha."""
    if lingua == "EN":
        return (template.get(f"{chave}_en") or "").strip() or template[chave]
    return template[chave]


def render_template(template: dict, row: pd.Series, params: dict) -> tuple[str, str]:
    lingua = lingua_cliente(row)
    ctx = {
        "nome": row["Nome"],
        "nif": row["NIF"],
        "email": row["Email"],
        "pag1": formatar_valor(row["Pag1"]),
        "pag2": formatar_valor(row["Pag2"]),
        "pag3": formatar_valor(row["Pag3"]),
        "total": formatar_valor(row["Total_PPC"]),
        "data1": params["data1"].strftime("%d/%m/%Y"),
        "data2": params["data2"].strftime("%d/%m/%Y"),
        "data3": params["data3"].strftime("%d/%m/%Y"),
        "ano_dados": params.get("ano_dados", 2025),
        "ano_pagamentos": params.get("ano_pagamentos", 2026),
    }
    assunto = texto_template(template, "assunto", lingua).format(**ctx)
    corpo = texto_template(template, "corpo", lingua).format(**ctx)
    return assunto, corpo


def editor_template_bilingue(tpl: dict, prefixo_key: str, altura: int = 260):
    """Widget partilhado: edita o assunto/corpo de um template em PT e EN
    (dois separadores). Altera o dicionário 'tpl' diretamente."""
    tab_pt, tab_en = st.tabs(["🇵🇹 Português", "🇬🇧 English"])
    with tab_pt:
        tpl["assunto"] = st.text_input("Assunto (PT)", value=tpl.get("assunto", ""), key=f"{prefixo_key}_assunto_pt")
        tpl["corpo"] = st.text_area("Corpo (PT)", value=tpl.get("corpo", ""), height=altura, key=f"{prefixo_key}_corpo_pt")
    with tab_en:
        tpl["assunto_en"] = st.text_input("Assunto (EN)", value=tpl.get("assunto_en", ""), key=f"{prefixo_key}_assunto_en")
        tpl["corpo_en"] = st.text_area("Corpo (EN)", value=tpl.get("corpo_en", ""), height=altura, key=f"{prefixo_key}_corpo_en")
        st.caption("Se deixares a versão EN vazia, os clientes EN recebem a versão PT.")


def enviar_email(smtp_cfg, destinatario, assunto, corpo, anexos, cc=None, assinatura_html=""):
    """Envia o email em texto E em HTML (multipart/alternative): quem abre num
    cliente moderno vê a versão HTML (com a assinatura da SERVE, se definida em
    Configurações); os restantes veem o texto simples."""
    import html as html_mod

    cc_list = [e.strip() for e in (cc or []) if e and e.strip()]

    msg = MIMEMultipart("mixed")
    msg["From"] = smtp_cfg["remetente"]
    msg["To"] = destinatario
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    msg["Subject"] = assunto

    corpo_html = (
        '<html><body><div style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #1a1a2e;">'
        + html_mod.escape(corpo).replace("\n", "<br>\n")
        + ("<br><br>" + assinatura_html if assinatura_html else "")
        + "</div></body></html>"
    )
    alternativa = MIMEMultipart("alternative")
    alternativa.attach(MIMEText(corpo, "plain", "utf-8"))
    alternativa.attach(MIMEText(corpo_html, "html", "utf-8"))
    msg.attach(alternativa)

    for filename, filebytes in anexos:
        part = MIMEApplication(filebytes, Name=filename)
        part["Content-Disposition"] = f'attachment; filename="{filename}"'
        msg.attach(part)

    todos_destinatarios = [destinatario] + cc_list

    context = ssl.create_default_context()
    if smtp_cfg["tls"]:
        with smtplib.SMTP(smtp_cfg["host"], smtp_cfg["porta"], timeout=30) as server:
            server.starttls(context=context)
            server.login(smtp_cfg["utilizador"], smtp_cfg["password"])
            server.sendmail(smtp_cfg["remetente"], todos_destinatarios, msg.as_string())
    else:
        with smtplib.SMTP_SSL(smtp_cfg["host"], smtp_cfg["porta"], context=context, timeout=30) as server:
            server.login(smtp_cfg["utilizador"], smtp_cfg["password"])
            server.sendmail(smtp_cfg["remetente"], todos_destinatarios, msg.as_string())


def extrair_nif_de_filename(filename: str):
    m = re.search(r"\b(\d{9})\b", filename)
    return m.group(1) if m else None


def validar_nif(nif: str) -> bool:
    """Valida o dígito de controlo de um NIF português (9 dígitos, módulo 11).
    Serve para apanhar NIFs mal escritos na importação/edição — um NIF inválido
    aqui é quase de certeza um erro de digitação."""
    nif = str(nif).strip()
    if not re.fullmatch(r"\d{9}", nif):
        return False
    soma = sum(int(d) * p for d, p in zip(nif[:8], range(9, 1, -1)))
    resto = soma % 11
    controlo = 0 if resto < 2 else 11 - resto
    return int(nif[8]) == controlo


def nifs_invalidos(df: pd.DataFrame) -> list:
    """Devolve a lista de NIFs presentes no df que falham a validação."""
    if "NIF" not in df.columns:
        return []
    return [n for n in df["NIF"].astype(str).str.strip() if n and not validar_nif(n)]


# ---------------------------------------------------------------------------
# Encriptação das passwords SMTP — usa a chave SMTP_ENC_KEY definida nos
# Secrets. Sem chave configurada, funciona na mesma mas guarda em texto simples
# (com aviso). Gerar uma chave:
#   python -c "from cryptography.fernet import Fernet; print(Fernet.generate_key().decode())"
# ---------------------------------------------------------------------------
def _fernet():
    chave = st.secrets.get("SMTP_ENC_KEY")
    if not chave:
        return None
    try:
        return Fernet(chave.encode() if isinstance(chave, str) else chave)
    except Exception:
        return None


def encriptar_password(password: str) -> str:
    f = _fernet()
    return f.encrypt(password.encode()).decode() if f else password


def desencriptar_password(guardada: str) -> str:
    """Desencripta a password guardada. Se não estiver encriptada (contas
    antigas, ou sem SMTP_ENC_KEY), devolve-a tal como está."""
    if not guardada:
        return guardada
    f = _fernet()
    if not f:
        return guardada
    try:
        return f.decrypt(guardada.encode()).decode()
    except (InvalidToken, Exception):
        return guardada  # legado: estava em texto simples


# ---------------------------------------------------------------------------
# Contas de email (SMTP) nomeadas e persistentes — cada gestor pode ter as suas
# próprias contas privadas, e o admin pode criar contas partilhadas (ex: "IRS
# Geral", "Segurança Social") que ficam visíveis para toda a equipa. Isto
# substitui ter de escrever utilizador/password de cada vez que se envia email.
# ---------------------------------------------------------------------------
def carregar_contas_email() -> list:
    client = get_client()
    resp = client.table("smtp_contas").select("*").order("nome").execute()
    return resp.data or []


def criar_conta_email(nome: str, host: str, porta: int, tls: bool, utilizador: str, password: str, remetente: str, partilhada: bool):
    client = get_client()
    client.table("smtp_contas").insert({
        "nome": nome, "host": host, "porta": porta, "tls": tls,
        "utilizador": utilizador, "password": encriptar_password(password), "remetente": remetente or utilizador,
        "proprietario_id": st.session_state.user.id, "partilhada": partilhada,
    }).execute()


def apagar_conta_email(conta_id: int):
    client = get_client()
    client.table("smtp_contas").delete().eq("id", conta_id).execute()


def escolher_conta_email(contexto: str) -> dict:
    """Widget reutilizável: escolher de entre as contas de email já guardadas
    (privadas + partilhadas), com atalho para criar uma nova ou apagar as tuas.
    'contexto' é só uma etiqueta (ex: 'ppc', 'irs') para lembrar a última
    escolhida em cada página, separadamente."""
    contas = carregar_contas_email()
    st.markdown("### Conta de Email")

    smtp_cfg = {"host": "", "porta": 587, "tls": True, "utilizador": "", "password": "", "remetente": ""}
    if contas:
        opcoes = {
            f"{c['nome']}" + (" 🌐 partilhada" if c["partilhada"] else " 🔒 privada"): c
            for c in contas
        }
        rotulos = list(opcoes.keys())
        chave_escolha = f"conta_email_escolhida_{contexto}"
        indice_default = rotulos.index(st.session_state[chave_escolha]) if st.session_state.get(chave_escolha) in rotulos else 0
        escolhida_label = st.selectbox("Enviar a partir de", rotulos, index=indice_default, key=f"select_{chave_escolha}")
        st.session_state[chave_escolha] = escolhida_label
        conta = opcoes[escolhida_label]
        smtp_cfg = {
            "host": conta["host"], "porta": int(conta["porta"]), "tls": conta["tls"],
            "utilizador": conta["utilizador"], "password": desencriptar_password(conta["password"]),
            "remetente": conta["remetente"] or conta["utilizador"],
        }
        if st.button("📨 Enviar email de teste", key=f"teste_conta_{contexto}",
                     help="Envia um email de teste para a própria conta, para confirmares que está bem configurada antes de a usares em envios em massa."):
            try:
                enviar_email(
                    smtp_cfg, smtp_cfg["remetente"], "Teste — Gestão Fiscal SERVE",
                    "Este é um email de teste enviado a partir da plataforma Gestão Fiscal SERVE.\n"
                    "Se o estás a ler, a conta está bem configurada.", [],
                )
                st.success(f"Email de teste enviado para {smtp_cfg['remetente']} — confirma a caixa de entrada.")
            except Exception as e:
                st.error(f"Falha no teste: {e}")
    else:
        st.info("Ainda não tens nenhuma conta de email guardada — cria uma abaixo.")

    with st.expander("➕ Adicionar / gerir contas de email"):
        with st.form(f"nova_conta_email_{contexto}", clear_on_submit=True):
            nome = st.text_input("Nome da conta (ex: 'IRS Geral', 'Miguel — pessoal')")
            host = st.text_input("Servidor SMTP", value="smtp.office365.com")
            c1, c2 = st.columns(2)
            with c1:
                utilizador = st.text_input("Utilizador (email de login)")
                porta = st.number_input("Porta", value=587, step=1)
            with c2:
                remetente = st.text_input("Remetente (From, opcional — usa o utilizador se vazio)")
                tls = st.checkbox("Usar STARTTLS (porta 587)", value=True)
            password = st.text_input("Password / App Password", type="password")
            partilhada = False
            if sou_admin():
                partilhada = st.checkbox("Tornar visível para toda a equipa (conta partilhada)")
            st.caption("Gmail: smtp.gmail.com, porta 587. Office365/Outlook: smtp.office365.com, porta 587. Usa sempre uma 'App Password', nunca a password principal da conta.")
            if not st.secrets.get("SMTP_ENC_KEY"):
                st.warning("⚠️ Sem SMTP_ENC_KEY configurada nos Secrets, as passwords ficam guardadas sem encriptação. Ver GUIA_SUPABASE.md (atualização v5).")
            if st.form_submit_button("💾 Guardar conta"):
                if nome and utilizador and password:
                    criar_conta_email(nome, host, int(porta), tls, utilizador, password, remetente, partilhada)
                    st.success(f"Conta '{nome}' guardada.")
                    st.rerun()
                else:
                    st.error("Preenche pelo menos nome, utilizador e password.")

        if contas:
            st.divider()
            st.caption("Contas existentes:")
            for c in contas:
                pode_apagar = c["proprietario_id"] == st.session_state.user.id or sou_admin()
                col_a, col_b = st.columns([5, 1])
                with col_a:
                    marca = "🌐 partilhada" if c["partilhada"] else "🔒 privada"
                    st.write(f"**{c['nome']}** — {c['utilizador']} ({marca})")
                with col_b:
                    if pode_apagar and st.button("🗑️ Apagar", key=f"apagar_conta_{c['id']}_{contexto}"):
                        apagar_conta_email(c["id"])
                        st.rerun()

    return smtp_cfg


# ---------------------------------------------------------------------------
# Supabase Storage — guias e faturas em PDF PERSISTENTES (bucket "guias").
# Carregas o PDF uma vez e ele fica guardado: já não se perde ao fechar o
# browser. Estrutura de pastas dentro do bucket:
#   ppc/<n_pagamento>/<nif>.pdf      — guias de PPC
#   irs/<nif>/guia.pdf               — guia de pagamento de IRS
#   irs/<nif>/fatura.pdf             — fatura do serviço de IRS
# ---------------------------------------------------------------------------
BUCKET_GUIAS = "guias"


def storage_upload_pdf(caminho: str, conteudo: bytes):
    """Carrega (ou substitui) um PDF no bucket de guias."""
    client = get_client()
    client.storage.from_(BUCKET_GUIAS).upload(
        caminho, conteudo, {"content-type": "application/pdf", "upsert": "true"}
    )


def storage_download_pdf(caminho: str):
    """Devolve os bytes do PDF, ou None se não existir."""
    try:
        return get_client().storage.from_(BUCKET_GUIAS).download(caminho)
    except Exception:
        return None


def storage_listar(pasta: str) -> set:
    """Nomes de ficheiro (sem o caminho) existentes numa pasta do bucket."""
    try:
        itens = get_client().storage.from_(BUCKET_GUIAS).list(pasta)
        return {i["name"] for i in (itens or []) if i.get("name")}
    except Exception:
        return set()


def storage_apagar(caminho: str):
    try:
        get_client().storage.from_(BUCKET_GUIAS).remove([caminho])
    except Exception:
        pass


# ---------------------------------------------------------------------------
# IRS — dados específicos (tabela própria, ligada por NIF) e leitura de PDFs
# ---------------------------------------------------------------------------
IRS_COLS = ["NIF", "Numero_Liquidacao", "Valor_Apurado", "Valor_Pendente", "Incluido_Avenca", "Email_Enviado"]
IRS_NUM_COLS = ["Valor_Apurado", "Valor_Pendente"]
IRS_BOOL_COLS = ["Incluido_Avenca", "Email_Enviado"]

IRS_COLUMN_MAP_TO_DB = {
    "NIF": "nif", "Numero_Liquidacao": "numero_liquidacao",
    "Valor_Apurado": "valor_apurado", "Valor_Pendente": "valor_pendente",
    "Incluido_Avenca": "incluido_avenca", "Email_Enviado": "email_enviado",
}
IRS_COLUMN_MAP_FROM_DB = {v: k for k, v in IRS_COLUMN_MAP_TO_DB.items()}

DEFAULT_TEMPLATE_IRS = {
    "assunto": "Liquidação de IRS — {nome}",
    "corpo": (
        "Exmo(a). Sr(a).,\n\n"
        "Junto enviamos a Demonstração de Liquidação de IRS referente ao ano de {ano_dados} "
        "(NIF {nif}{ref_liquidacao}).\n\n"
        "{frase_valor}\n\n"
        "{frase_pendente}"
        "Segue também em anexo a guia de pagamento, quando aplicável.\n\n"
        "Ficamos ao dispor para qualquer esclarecimento.\n\n"
        "Com os melhores cumprimentos,"
    ),
    "assunto_en": "Personal Income Tax (IRS) Assessment — {nome}",
    "corpo_en": (
        "Dear Sir or Madam,\n\n"
        "Please find attached the Personal Income Tax (IRS) Assessment Statement for the year {ano_dados} "
        "(tax no. {nif}{ref_liquidacao}).\n\n"
        "{frase_valor}\n\n"
        "{frase_pendente}"
        "The payment form is also attached, where applicable.\n\n"
        "We remain at your disposal for any clarification.\n\n"
        "Best regards,"
    ),
}


def clean_irs_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for c in IRS_COLS:
        if c not in df.columns:
            if c in IRS_BOOL_COLS:
                df[c] = False
            elif c in IRS_NUM_COLS:
                df[c] = 0.0
            else:
                df[c] = ""
    df["NIF"] = df["NIF"].fillna("").astype(str).str.strip()
    df["Numero_Liquidacao"] = df["Numero_Liquidacao"].fillna("").astype(str).str.strip()
    for c in IRS_NUM_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    for c in IRS_BOOL_COLS:
        df[c] = df[c].fillna(False).astype(bool)
    df = df[df["NIF"] != ""]
    return df[IRS_COLS]


def carregar_irs_db() -> pd.DataFrame:
    client = get_client()
    resp = client.table("irs_dados").select("*").execute()
    rows = resp.data or []
    if not rows:
        return pd.DataFrame(columns=IRS_COLS)
    df = pd.DataFrame(rows).rename(columns=IRS_COLUMN_MAP_FROM_DB)
    return clean_irs_df(df)


def guardar_irs_db(df: pd.DataFrame, nifs_antes: set = None):
    """Grava por diferença (upsert + apagar só os que desapareceram) — ver nota
    em guardar_clientes_db."""
    client = get_client()
    df2 = clean_irs_df(df).copy()
    if nifs_antes is None:
        antes = clean_irs_df(st.session_state.get("irs_dados", pd.DataFrame(columns=IRS_COLS)))
        nifs_antes = set(antes["NIF"])
    para_apagar = nifs_antes - set(df2["NIF"])
    if para_apagar:
        client.table("irs_dados").delete().in_("nif", list(para_apagar)).execute()
    if not df2.empty:
        registos = df2.rename(columns=IRS_COLUMN_MAP_TO_DB).to_dict("records")
        client.table("irs_dados").upsert(registos, on_conflict="nif").execute()


def persistir_irs(df: pd.DataFrame):
    df = clean_irs_df(df)
    guardar_irs_db(df)
    st.session_state.irs_dados = df


def montar_base_irs() -> pd.DataFrame:
    """Junta os clientes com 'Aplica_IRS' ligado com os respetivos dados de IRS
    (mesmo que ainda não tenham nenhum valor preenchido)."""
    clientes = clean_clientes_df(st.session_state.clientes)
    elegiveis = clientes[clientes["Aplica_IRS"]].copy()
    irs = clean_irs_df(st.session_state.get("irs_dados", pd.DataFrame(columns=IRS_COLS)))
    base = elegiveis.merge(irs, on="NIF", how="left")
    for c in IRS_NUM_COLS:
        base[c] = pd.to_numeric(base[c], errors="coerce").fillna(0.0)
    for c in IRS_BOOL_COLS:
        base[c] = base[c].fillna(False).astype(bool)
    base["Numero_Liquidacao"] = base["Numero_Liquidacao"].fillna("")
    return base


def _parse_valor_pt(texto: str):
    """Converte um número em formato português (ex: '-1.234,56') encontrado em
    texto extraído de PDF para float. Devolve None se não conseguir converter."""
    if texto is None:
        return None
    s = texto.strip().replace(" ", "").replace("€", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _extrair_texto_pdf(ficheiro_bytes: bytes) -> str:
    texto_paginas = []
    with pdfplumber.open(io.BytesIO(ficheiro_bytes)) as pdf:
        for pagina in pdf.pages:
            texto_paginas.append(pagina.extract_text() or "")
    return "\n".join(texto_paginas)


def extrair_dados_liquidacao_irs(ficheiro_bytes: bytes, nif_esperado: str = None) -> dict:
    """Lê uma 'Demonstração de Liquidação de IRS' da Autoridade Tributária e tenta
    extrair o número de liquidação e o valor final (a pagar, a receber, ou
    apurado quando é zero). Se 'nif_esperado' for indicado, verifica apenas se
    esse NIF aparece em algum lado do documento — não tenta identificar/listar
    todos os NIFs do documento (podem ser um ou dois, em declarações conjuntas,
    e não interessa a posição, só se o do cliente lá está). Devolve sempre um
    dicionário com as chaves — valores a None se não encontrar, para a página
    poder avisar e pedir preenchimento manual em vez de adivinhar."""
    resultado = {
        "nif_confirmado": None, "numero_liquidacao": None, "periodo": None,
        "valor_apurado": None, "tipo_valor": None, "texto_bruto": "",
    }
    try:
        texto = _extrair_texto_pdf(ficheiro_bytes)
    except Exception:
        return resultado
    resultado["texto_bruto"] = texto

    if nif_esperado:
        resultado["nif_confirmado"] = bool(re.search(rf"\b{re.escape(nif_esperado)}\b", texto))

    # Número de liquidação (formato "AAAA.dígitos") seguido do período de
    # rendimentos — não precisamos de capturar o(s) NIF(s) que vêm antes.
    m_linha = re.search(
        r"(?P<numero>\d{4}\.\d+)\s+(?P<periodo_ini>\d{4}-\d{2}-\d{2})\s+a\s+(?P<periodo_fim>\d{4}-\d{2}-\d{2})",
        texto,
    )
    if m_linha:
        resultado["numero_liquidacao"] = m_linha.group("numero")
        resultado["periodo"] = f"{m_linha.group('periodo_ini')} a {m_linha.group('periodo_fim')}"

    # O valor final vem sempre rotulado — "Valor a pagar", "Valor a receber" ou
    # "Valor apurado" (quando não há nada a pagar nem a receber). O próprio
    # rótulo diz-nos o sinal, por isso não precisamos de adivinhar a partir do
    # número (o número no PDF vem sempre positivo, mesmo quando é um reembolso).
    m_valor = re.search(r"Valor (a pagar|a receber|apurado)\s+(-?[\d\.]+,\d{2})", texto)
    if m_valor:
        rotulo = m_valor.group(1)
        numero = _parse_valor_pt(m_valor.group(2))
        if numero is not None:
            if rotulo == "a receber":
                numero = -abs(numero)
            elif rotulo == "a pagar":
                numero = abs(numero)
        resultado["valor_apurado"] = numero
        resultado["tipo_valor"] = rotulo

    return resultado


def extrair_dados_pendentes_irs(ficheiro_bytes: bytes) -> dict:
    """Lê um 'Controlo de Pendentes' (extrato de faturas em dívida à SERVE, gerado
    pelo TOConline) e tenta extrair o NIF do cliente e o total pendente."""
    resultado = {"nif": None, "valor_pendente": None, "texto_bruto": ""}
    try:
        texto = _extrair_texto_pdf(ficheiro_bytes)
    except Exception:
        return resultado
    resultado["texto_bruto"] = texto

    m_nif = re.search(r"EUR\s+(\d{9})", texto)
    if m_nif:
        resultado["nif"] = m_nif.group(1)

    m_total = re.search(r"TOTAL PENDENTE\s+([\d\.,]+)", texto)
    if m_total:
        resultado["valor_pendente"] = _parse_valor_pt(m_total.group(1))

    return resultado


def render_template_irs(template: dict, row: pd.Series) -> tuple[str, str]:
    lingua = lingua_cliente(row)
    valor = row.get("Valor_Apurado", 0.0) or 0.0
    if lingua == "EN":
        if valor > 0:
            frase_valor = f"The assessment results in an amount payable of {formatar_valor(valor)} €."
        elif valor < 0:
            frase_valor = f"The assessment results in a refund of {formatar_valor(abs(valor))} €."
        else:
            frase_valor = "The assessment results in no amount payable or refundable."
    else:
        if valor > 0:
            frase_valor = f"Do apuramento efetuado, resulta um valor a pagar de {formatar_valor(valor)} €."
        elif valor < 0:
            frase_valor = f"Do apuramento efetuado, resulta um valor a receber (reembolso) de {formatar_valor(abs(valor))} €."
        else:
            frase_valor = "Do apuramento efetuado, não resulta qualquer valor a pagar ou a receber."

    pendente = row.get("Valor_Pendente", 0.0) or 0.0
    if pendente > 0:
        if lingua == "EN":
            frase_pendente = (
                f"We would also like to inform you that, according to our records, an amount of "
                f"{formatar_valor(pendente)} € remains outstanding to SERVE for professional fees.\n\n"
            )
        else:
            frase_pendente = (
                f"Informamos ainda que, de acordo com os nossos registos, tem pendente o valor de "
                f"{formatar_valor(pendente)} € referente a honorários em dívida à SERVE.\n\n"
            )
    else:
        frase_pendente = ""

    if row.get("Numero_Liquidacao"):
        ref_liquidacao = (f", assessment no. {row['Numero_Liquidacao']}" if lingua == "EN"
                          else f", n.º de liquidação {row['Numero_Liquidacao']}")
    else:
        ref_liquidacao = ""

    params = st.session_state.get("params", {})
    ctx = {
        "nome": row["Nome"],
        "nif": row["NIF"],
        "email": row["Email"],
        "ref_liquidacao": ref_liquidacao,
        "frase_valor": frase_valor,
        "frase_pendente": frase_pendente,
        "ano_dados": params.get("ano_dados", 2025),
        "ano_pagamentos": params.get("ano_pagamentos", 2026),
    }
    assunto = texto_template(template, "assunto", lingua).format(**ctx)
    corpo = texto_template(template, "corpo", lingua).format(**ctx)
    return assunto, corpo


def gerar_excel_irs(base_irs: pd.DataFrame, params: dict) -> bytes:
    """Folha de controlo das liquidações de IRS (mesmo estilo do Excel do PPC)."""
    ano_dados = params.get("ano_dados", 2025)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Controlo IRS {ano_dados}"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    FONT = "Arial"
    HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    ENVIADO_FILL = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    HEADER_FONT = Font(name=FONT, color="FFFFFF", bold=True, size=10)
    TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
    BLACK = Font(name=FONT, color="000000")
    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"Controlo de Liquidações de IRS — {ano_dados}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    headers = [
        "N.º", "NIF", "Nome", "Email",
        "Nº Liquidação", "Valor Apurado (€)", "Pendente à SERVE (€)",
        "Incluído na Avença", "Email Enviado", "Só IRS (avulso)", "Notas",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[4].height = 30

    widths = [9, 12, 26, 24, 16, 15, 15, 12, 12, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 5
    for _, r in base_irs.iterrows():
        vals = [
            r.get("Numero_Cliente", ""), r["NIF"], r["Nome"], r["Email"],
            r["Numero_Liquidacao"], r["Valor_Apurado"], r["Valor_Pendente"],
            "Sim" if r["Incluido_Avenca"] else "Não",
            "Sim" if r["Email_Enviado"] else "Não",
            "Sim" if r.get("IRS_Avulso", False) else "Não",
            r.get("Notas", ""),
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = BLACK
            c.border = BORDER
            if i in (6, 7):
                c.number_format = "#,##0.00"
        if r["Email_Enviado"]:
            for i in range(1, len(vals) + 1):
                ws.cell(row=row, column=i).fill = ENVIADO_FILL
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Segurança Social (DMR/DRI) — envio mensal de declarações e guias.
# Estado "email enviado" guardado por cliente e por mês na tabela ss_dados.
# Documentos no Storage: ss/<mes>/guia/<nif>.pdf, ss/<mes>/dmr/<nif>.pdf e
# extras em ss/<mes>/extra/<nif>__<nome do ficheiro>.
# ---------------------------------------------------------------------------
MESES_PT = ["janeiro", "fevereiro", "março", "abril", "maio", "junho",
            "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
MESES_EN = ["January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"]

DEFAULT_TEMPLATE_SS = {
    "assunto": "Segurança Social — {mes_nome} — {nome}",
    "corpo": (
        "Exmo(a). Sr(a).,\n\n"
        "Junto enviamos a documentação da Segurança Social referente a {mes_nome}: {lista_docs}.\n\n"
        "O pagamento deverá ser efetuado até {data_limite}.\n\n"
        "Ficamos ao dispor para qualquer esclarecimento.\n\n"
        "Com os melhores cumprimentos,"
    ),
    "assunto_en": "Social Security — {mes_nome} — {nome}",
    "corpo_en": (
        "Dear Sir or Madam,\n\n"
        "Please find attached the Social Security documentation for {mes_nome}: {lista_docs}.\n\n"
        "Payment should be made by {data_limite}.\n\n"
        "We remain at your disposal for any clarification.\n\n"
        "Best regards,"
    ),
}


def nome_mes(mes: str, lingua: str = "PT") -> str:
    """'2026-06' -> 'junho de 2026' (PT) ou 'June 2026' (EN)."""
    ano, m = mes.split("-")
    m = int(m)
    return f"{MESES_EN[m - 1]} {ano}" if lingua == "EN" else f"{MESES_PT[m - 1]} de {ano}"


def data_limite_ss(mes: str) -> date:
    """Dia 25 do mês seguinte ao mês de referência (remunerações de junho ->
    pagamento até 25 de julho)."""
    ano, m = (int(x) for x in mes.split("-"))
    m += 1
    if m == 13:
        m, ano = 1, ano + 1
    return date(ano, m, 25)


def lista_meses_ss(quantos: int = 18) -> list:
    """Meses de referência disponíveis no seletor, do mais recente para trás,
    a começar no mês atual."""
    hoje = date.today()
    ano, m = hoje.year, hoje.month
    meses = []
    for _ in range(quantos):
        meses.append(f"{ano:04d}-{m:02d}")
        m -= 1
        if m == 0:
            m, ano = 12, ano - 1
    return meses


def carregar_ss_mes_db(mes: str) -> dict:
    """Estado 'email enviado' de todos os clientes num mês: {nif: True/False}."""
    try:
        resp = get_client().table("ss_dados").select("nif, email_enviado").eq("mes", mes).execute()
        return {r["nif"]: bool(r["email_enviado"]) for r in (resp.data or [])}
    except Exception:
        return {}  # tabela ainda não criada (v7 por correr)


def marcar_ss_enviado_db(nif: str, mes: str, enviado: bool = True):
    get_client().table("ss_dados").upsert(
        {"nif": nif, "mes": mes, "email_enviado": enviado}, on_conflict="nif,mes"
    ).execute()


def montar_base_ss() -> pd.DataFrame:
    """Clientes com 'Aplica SS' ligado (o estado por mês junta-se na página)."""
    clientes = clean_clientes_df(st.session_state.clientes)
    return clientes[clientes["Aplica_SS"]].copy()


def docs_ss_cliente(mes: str, nif: str, guias_set: set, dmrs_set: set, extras_dict: dict) -> list:
    """Lista dos documentos disponíveis para um cliente neste mês, a partir dos
    conjuntos já lidos do Storage (para não fazer chamadas a mais):
    devolve ex: ["guia", "dmr", "extra:Recibo.pdf"]."""
    docs = []
    if nif in guias_set:
        docs.append("guia")
    if nif in dmrs_set:
        docs.append("dmr")
    for nome_extra in extras_dict.get(nif, []):
        docs.append(f"extra:{nome_extra}")
    return docs


def listar_extras_ss(mes: str) -> dict:
    """Extras carregados no mês, agrupados por NIF: {nif: [nome1.pdf, ...]}."""
    extras = {}
    for nome in storage_listar(f"ss/{mes}/extra"):
        if "__" in nome:
            nif, nome_ficheiro = nome.split("__", 1)
            extras.setdefault(nif, []).append(nome_ficheiro)
    return extras


def render_template_ss(template: dict, row: pd.Series, mes: str, docs: list) -> tuple[str, str]:
    """Monta o email da Segurança Social na língua do cliente. 'docs' é a lista
    devolvida por docs_ss_cliente — usada para escrever a frase dos anexos."""
    lingua = lingua_cliente(row)
    partes = []
    for d in docs:
        if d == "guia":
            partes.append("payment form" if lingua == "EN" else "guia de pagamento")
        elif d == "dmr":
            partes.append("DMR")
        elif d.startswith("extra:"):
            partes.append(d.split(":", 1)[1])
    if partes:
        lista_docs = ", ".join(partes)
    else:
        lista_docs = "the attached documents" if lingua == "EN" else "os documentos em anexo"

    ctx = {
        "nome": row["Nome"],
        "nif": row["NIF"],
        "email": row["Email"],
        "mes_nome": nome_mes(mes, lingua),
        "data_limite": data_limite_ss(mes).strftime("%d/%m/%Y"),
        "lista_docs": lista_docs,
    }
    assunto = texto_template(template, "assunto", lingua).format(**ctx)
    corpo = texto_template(template, "corpo", lingua).format(**ctx)
    return assunto, corpo
